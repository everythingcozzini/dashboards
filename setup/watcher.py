#!/usr/bin/env python3
"""
Cozzini Dashboard Auto-Updater
Watches for new Excel files, parses them, updates dashboard HTML, and pushes to GitHub.

Usage:
    python3 watcher.py          # Run the live file watcher
    python3 watcher.py --once   # Process all pending files once and exit

File naming convention (files are matched by keyword, not strict pattern):
    *nps*existing*.xlsx         → nps.html
    *customer*churn*.xlsx       → customerchurn.html
    *product*churn*.xlsx        → productchurn.html
    *ces_price*.xlsx            → ces.html (pricing)
    *ces_onboard*.xlsx          → ces.html (onboarding)
    *ces_knife*.xlsx            → ces.html (knife sharpness)
    *ces_driver_service*.xlsx   → ces.html (driver service)
    *ces_invoice_payment*.xlsx  → ces.html (invoice payment)
    *ces_invoice_under*.xlsx    → ces.html (invoice understanding)
    *ces_driver_last*.xlsx      → ces.html (driver last visit - currently unused)
"""

import os
import sys
import re
import json
import time
import logging
import subprocess
from datetime import datetime
from pathlib import Path
from collections import defaultdict

import openpyxl
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DASH_DIR = Path(__file__).resolve().parent
VOC_DIR = DASH_DIR / "voc"   # Post-restructure: dashboard HTMLs live under voc/
LOG_FILE = DASH_DIR / "watcher.log"
PROCESSED_LOG = DASH_DIR / ".processed_files.json"
AM_CACHE = DASH_DIR / ".am_reference.json"  # DSD → [AM names] lookup for NPS drilldown

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger("dash-watcher")


def load_processed():
    if PROCESSED_LOG.exists():
        return json.loads(PROCESSED_LOG.read_text())
    return {}


def save_processed(data):
    PROCESSED_LOG.write_text(json.dumps(data, indent=2))


# ---------------------------------------------------------------------------
# File classification
# ---------------------------------------------------------------------------
def classify_file(filename):
    """Return (dashboard_type, ces_subtype) or (None, None) if not recognized."""
    fn = filename.lower()
    if not fn.endswith((".xlsx", ".xls")):
        return None, None

    # Area Manager reference file — caches DSD→AM mapping for NPS drilldown
    if "dsd" in fn and "am" in fn:
        return "am_reference", None

    if "nps" in fn and "existing" in fn:
        return "nps", None
    if "nps" in fn and "new" in fn:
        return "nps_new", None
    if "customer" in fn and "churn" in fn:
        return "customer_churn", None
    if "product" in fn and "churn" in fn:
        return "product_churn", None
    if "ces" in fn:
        if "price" in fn:
            return "ces", "price"
        if "onboard" in fn:
            return "ces", "onboarding"
        if "knife" in fn:
            return "ces", "knife"
        if "driver" in fn and "last" in fn:
            return "ces", "driver_last"
        if "driver" in fn:
            return "ces", "driver"
        if "invoice" in fn and "understand" in fn:
            return "ces", "invoice_understanding"
        if "invoice" in fn:
            return "ces", "invoice_payment"
    return None, None


def parse_am_reference(filepath):
    """Parse the DSD/AM reference spreadsheet and cache DSD→AM mapping to JSON.

    Expected columns (row 1 headers):
        DSD Name | Position | Location | Email | AM Name | Position | Location | Email

    Source uses 'First Last' format; NPS data uses 'Last, First'. We normalize
    DSD names to 'Last, First' so the NPS parser can look them up directly.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]

    dsd_to_ams = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        dsd_raw = ws.cell(r, 1).value
        am_raw = ws.cell(r, 5).value
        if not dsd_raw or not am_raw:
            continue
        dsd = str(dsd_raw).strip()
        # Normalize "First Last" → "Last, First"
        parts = dsd.rsplit(" ", 1)
        dsd_norm = f"{parts[1]}, {parts[0]}" if len(parts) == 2 else dsd
        am_name = str(am_raw).strip()
        if am_name not in dsd_to_ams[dsd_norm]:
            dsd_to_ams[dsd_norm].append(am_name)

    cache = {
        "updated": datetime.now().isoformat(timespec="seconds"),
        "source_file": os.path.basename(filepath),
        "dsd_to_ams": dict(dsd_to_ams),
    }
    AM_CACHE.write_text(json.dumps(cache, indent=2))
    log.info(f"  Cached AM reference — {len(dsd_to_ams)} DSDs, {sum(len(v) for v in dsd_to_ams.values())} AM links")
    return cache


def load_am_reference():
    """Return cached DSD→[AM names] dict, or empty dict if no cache exists."""
    if not AM_CACHE.exists():
        return {}
    try:
        return json.loads(AM_CACHE.read_text()).get("dsd_to_ams", {})
    except Exception as e:
        log.warning(f"Could not read AM cache: {e}")
        return {}


# ===========================================================================
# PARSERS — one per dashboard
# ===========================================================================

# ---------------------------------------------------------------------------
# NPS Parser
# ---------------------------------------------------------------------------
def parse_nps(filepath):
    """Parse NPS Excel → dict with all data needed for nps.html.

    Supports two Excel formats:
      - NEW format: single 'Responses' sheet with raw per-respondent rows
        (Sharpening Center | ... | Date/Time | NPS score | ... | DSD | NPS)
      - OLD format: multi-sheet ('NPS E How Likely', 'Chart', 'Drilldown')
        pre-aggregated pivot export

    Both produce the same return shape. AM drilldown is only populated from
    the OLD format (new export does not include Area Manager).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheetnames = wb.sheetnames

    # ----------------------------------------------------------------------
    # NEW FORMAT: single 'Responses' sheet
    # ----------------------------------------------------------------------
    if "Responses" in sheetnames and "NPS E How Likely" not in sheetnames:
        ws = wb["Responses"]

        # Auto-detect score column by header (in case columns shift)
        headers = {str(c.value).strip(): c.column for c in ws[1] if c.value}
        col_center = headers.get("Sharpening Center", 1)
        col_date = headers.get("Date/Time", 6)
        # Score column: prefer col with "recommend" question, fallback to 7
        col_score = 7
        for h, c in headers.items():
            if "recommend" in h.lower() or h.strip() == "NPS":
                # Take the first "recommend" match (raw score), not the "NPS" label column
                if "recommend" in h.lower():
                    col_score = c
                    break
        col_dsd = headers.get("DSD", 12)

        rows = []
        for r in range(2, ws.max_row + 1):
            center = ws.cell(r, col_center).value
            score = ws.cell(r, col_score).value
            dsd = ws.cell(r, col_dsd).value
            dt = ws.cell(r, col_date).value
            if score is not None:
                try:
                    score_int = int(score)
                except (ValueError, TypeError):
                    continue
                if 0 <= score_int <= 10:
                    rows.append({
                        "center": str(center).strip() if center else "Unassigned",
                        "score": score_int,
                        "dsd": str(dsd).strip() if dsd else "",
                        "date": dt if dt else "",
                    })

        if not rows:
            return None

        total = len(rows)
        scores = [r["score"] for r in rows]
        avg_score = sum(scores) / total

        detractors = sum(1 for s in scores if s <= 6)
        passives = sum(1 for s in scores if s in (7, 8))
        promoters = sum(1 for s in scores if s >= 9)
        nps = round((promoters / total - detractors / total) * 100, 1)

        # Score distribution 1-10 (matches old dashboard buckets)
        dist = [sum(1 for s in scores if s == i) for i in range(1, 11)]

        # Date range
        dates = []
        for r in rows:
            d = r["date"]
            if hasattr(d, "strftime"):
                dates.append(d)
            elif isinstance(d, str) and d:
                for fmt in ["%m/%d/%Y %I:%M:%S %p", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y"]:
                    try:
                        clean = d.split(" CST")[0].split(" CDT")[0].split(" EST")[0].split(" EDT")[0].strip()
                        dates.append(datetime.strptime(clean, fmt))
                        break
                    except ValueError:
                        continue
        date_min = min(dates).strftime("%b %d, %Y") if dates else "N/A"
        date_max = max(dates).strftime("%b %d, %Y") if dates else "N/A"

        # Aggregate: NPS by center
        by_center = {}
        for r in rows:
            by_center.setdefault(r["center"], []).append(r["score"])
        center_data = []
        drill_center = []
        for name, slist in by_center.items():
            n = len(slist)
            pro = sum(1 for s in slist if s >= 9)
            pas = sum(1 for s in slist if s in (7, 8))
            det = sum(1 for s in slist if s <= 6)
            c_nps = round((pro - det) / n * 100, 1)
            c_avg = sum(slist) / n
            center_data.append({"name": name, "nps": c_nps, "n": n})
            drill_center.append({
                "center": name,
                "det": det, "pas": pas, "pro": pro, "n": n,
                "proP": f"{pro/n*100:.1f}", "detP": f"{det/n*100:.1f}",
                "nps": f"{c_nps:.1f}", "avg": f"{c_avg:.2f}",
            })
        center_data.sort(key=lambda x: x["nps"], reverse=True)
        drill_center.sort(key=lambda x: float(x["nps"]), reverse=True)

        # Aggregate: Center + DSD
        by_dsd = {}
        for r in rows:
            key = (r["center"], r["dsd"])
            by_dsd.setdefault(key, []).append(r["score"])
        drill_dsd = []
        for (center, dsd), slist in by_dsd.items():
            n = len(slist)
            pro = sum(1 for s in slist if s >= 9)
            pas = sum(1 for s in slist if s in (7, 8))
            det = sum(1 for s in slist if s <= 6)
            d_nps = round((pro - det) / n * 100, 1)
            d_avg = sum(slist) / n
            drill_dsd.append({
                "center": center, "dsd": dsd if dsd else "(unassigned)",
                "det": det, "pas": pas, "pro": pro, "n": n,
                "nps": f"{d_nps:.1f}", "avg": f"{d_avg:.2f}",
            })
        drill_dsd.sort(key=lambda x: float(x["nps"]), reverse=True)

        # AM drilldown: build from cached DSD→AM reference if available.
        # Since NPS data has no AM-level granularity, each row represents the
        # DSD-level stats with the AM list (last names) rolled into one cell.
        dsd_to_ams = load_am_reference()
        drill_am = None
        if dsd_to_ams:
            drill_am = []
            for (center, dsd), slist in by_dsd.items():
                if not dsd:
                    continue  # skip unassigned DSDs — no AMs to attach
                n = len(slist)
                pro = sum(1 for s in slist if s >= 9)
                pas = sum(1 for s in slist if s in (7, 8))
                det = sum(1 for s in slist if s <= 6)
                d_nps = round((pro - det) / n * 100, 1)
                d_avg = sum(slist) / n
                ams_full = dsd_to_ams.get(dsd, [])
                # Display as last names only, sorted, for readability.
                # Strip suffixes (Jr., Sr., II, III, IV) and trailing commas so
                # 'Kenneth Cochran, Jr.' → 'Cochran' not 'Jr.'
                ams_last = set()
                for a in ams_full:
                    clean = a.replace(",", "").strip()
                    tokens = [t for t in clean.split() if t.rstrip(".") not in ("Jr", "Sr", "II", "III", "IV")]
                    if tokens:
                        ams_last.add(tokens[-1])
                ams_display = ", ".join(sorted(ams_last)) if ams_last else "(not mapped)"
                drill_am.append({
                    "center": center,
                    "dsd": dsd,
                    "am": ams_display,
                    "det": det, "pas": pas, "pro": pro, "n": n,
                    "nps": f"{d_nps:.1f}", "avg": f"{d_avg:.2f}",
                })
            drill_am.sort(key=lambda x: float(x["nps"]), reverse=True)

        return {
            "total": total,
            "nps": nps,
            "avg_score": round(avg_score, 2),
            "promoters": promoters,
            "passives": passives,
            "detractors": detractors,
            "dist": dist,
            "center_data": center_data,
            "drill_center": drill_center,
            "drill_dsd": drill_dsd,
            "drill_am": drill_am,  # None if no AM ref cache → preserve existing HTML
            "date_min": date_min,
            "date_max": date_max,
        }

    # ----------------------------------------------------------------------
    # OLD FORMAT: multi-sheet pre-aggregated export
    # ----------------------------------------------------------------------
    ws = wb["NPS E How Likely"]
    rows = []
    for r in range(2, ws.max_row + 1):
        center = ws.cell(r, 1).value
        score = ws.cell(r, 7).value
        dsd = ws.cell(r, 12).value
        dt = ws.cell(r, 6).value
        if score is not None:
            rows.append({"center": str(center).strip() if center else "Unassigned", "score": int(score), "dsd": str(dsd).strip() if dsd else "", "date": str(dt) if dt else ""})

    if not rows:
        return None

    total = len(rows)
    scores = [r["score"] for r in rows]
    avg_score = sum(scores) / total

    # Segment counts
    detractors = sum(1 for s in scores if s <= 6)
    passives = sum(1 for s in scores if s in (7, 8))
    promoters = sum(1 for s in scores if s >= 9)
    nps = round((promoters / total - detractors / total) * 100, 1)

    # Score distribution (1-10)
    dist = [sum(1 for s in scores if s == i) for i in range(1, 11)]

    # Date range
    dates = []
    for r in rows:
        try:
            d = r["date"]
            if hasattr(d, "strftime"):
                dates.append(d)
            else:
                for fmt in ["%m/%d/%Y %I:%M:%S %p %Z", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y"]:
                    try:
                        dates.append(datetime.strptime(d.split(" CST")[0].split(" CDT")[0].strip(), fmt.replace(" %Z", "")))
                        break
                    except ValueError:
                        continue
        except Exception:
            pass

    date_min = min(dates).strftime("%b %d, %Y") if dates else "N/A"
    date_max = max(dates).strftime("%b %d, %Y") if dates else "N/A"

    # --- Chart sheet (NPS by center) ---
    ws_chart = wb["Chart"]
    center_data = []
    for r in range(2, ws_chart.max_row + 1):
        name = ws_chart.cell(r, 1).value
        n = ws_chart.cell(r, 2).value
        avg = ws_chart.cell(r, 3).value
        nps_val = ws_chart.cell(r, 4).value
        if name and n:
            center_data.append({"name": str(name).strip(), "nps": round(float(nps_val), 1), "n": int(n)})
    center_data.sort(key=lambda x: x["nps"], reverse=True)

    # --- Drilldown sheet ---
    ws_drill = wb["Drilldown"]

    # Center-level drilldown (cols A-I, starting row 3)
    drill_center = []
    for r in range(3, ws_drill.max_row + 1):
        name = ws_drill.cell(r, 1).value
        if not name or str(name).strip() == "":
            break
        det = ws_drill.cell(r, 2).value or 0
        pas = ws_drill.cell(r, 3).value or 0
        pro = ws_drill.cell(r, 4).value or 0
        n = ws_drill.cell(r, 5).value or 0
        proP = ws_drill.cell(r, 6).value or 0
        detP = ws_drill.cell(r, 7).value or 0
        nps_calc = ws_drill.cell(r, 8).value or 0
        avg_s = ws_drill.cell(r, 9).value or 0
        drill_center.append({
            "center": str(name).strip(),
            "det": int(det), "pas": int(pas), "pro": int(pro), "n": int(n),
            "proP": f"{float(proP):.1f}", "detP": f"{float(detP):.1f}",
            "nps": f"{float(nps_calc):.1f}", "avg": f"{float(avg_s):.2f}"
        })
    drill_center.sort(key=lambda x: float(x["nps"]), reverse=True)

    # DSD-level drilldown (cols K-T, starting row 3)
    drill_dsd = []
    for r in range(3, ws_drill.max_row + 1):
        name = ws_drill.cell(r, 11).value
        if not name or str(name).strip() == "":
            break
        dsd = ws_drill.cell(r, 12).value or ""
        det = ws_drill.cell(r, 13).value or 0
        pas = ws_drill.cell(r, 14).value or 0
        pro = ws_drill.cell(r, 15).value or 0
        n = ws_drill.cell(r, 16).value or 0
        nps_calc = ws_drill.cell(r, 19).value or 0
        avg_s = ws_drill.cell(r, 20).value or 0
        drill_dsd.append({
            "center": str(name).strip(), "dsd": str(dsd).strip(),
            "det": int(det), "pas": int(pas), "pro": int(pro), "n": int(n),
            "nps": f"{float(nps_calc):.1f}", "avg": f"{float(avg_s):.2f}"
        })
    drill_dsd.sort(key=lambda x: float(x["nps"]), reverse=True)

    # AM-level drilldown (cols V-AF, starting row 3)
    drill_am = []
    for r in range(3, ws_drill.max_row + 1):
        name = ws_drill.cell(r, 22).value
        if not name or str(name).strip() == "":
            break
        dsd = ws_drill.cell(r, 23).value or ""
        am = ws_drill.cell(r, 24).value or ""
        det = ws_drill.cell(r, 25).value or 0
        pas = ws_drill.cell(r, 26).value or 0
        pro = ws_drill.cell(r, 27).value or 0
        n = ws_drill.cell(r, 28).value or 0
        nps_calc = ws_drill.cell(r, 31).value or 0
        avg_s = ws_drill.cell(r, 32).value or 0
        drill_am.append({
            "center": str(name).strip(), "dsd": str(dsd).strip(), "am": str(am).strip(),
            "det": int(det), "pas": int(pas), "pro": int(pro), "n": int(n),
            "nps": f"{float(nps_calc):.1f}", "avg": f"{float(avg_s):.2f}"
        })
    drill_am.sort(key=lambda x: float(x["nps"]), reverse=True)

    return {
        "total": total,
        "nps": nps,
        "avg_score": round(avg_score, 2),
        "promoters": promoters,
        "passives": passives,
        "detractors": detractors,
        "dist": dist,
        "center_data": center_data,
        "drill_center": drill_center,
        "drill_dsd": drill_dsd,
        "drill_am": drill_am,
        "date_min": date_min,
        "date_max": date_max,
    }


# ---------------------------------------------------------------------------
# NPS New Customers Parser  (updates the bottom section of nps.html only)
# ---------------------------------------------------------------------------
def parse_nps_new_customers(filepath):
    """Parse nps_new_customers.xlsx → dict for the 'New Customer NPS' section of nps.html.

    Expected columns (row 1 headers):
        col 1  Sharpening Center
        col 7  Score (1-10)
        col 11 Company Name
        col 12 DSD (format 'Last, First')
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]

    respondents = []
    dist = [0] * 10           # index 0 = score 1, ..., index 9 = score 10
    center_counts = defaultdict(int)

    for r in range(2, ws.max_row + 1):
        score_raw = ws.cell(r, 7).value
        if score_raw is None or score_raw == "":
            continue
        try:
            score = int(score_raw)
        except (ValueError, TypeError):
            continue
        if not 1 <= score <= 10:
            continue

        center = ws.cell(r, 1).value
        company = ws.cell(r, 11).value
        dsd = ws.cell(r, 12).value

        center_s = str(center).strip() if center else "Unassigned"
        company_s = str(company).strip() if company else "—"
        dsd_s = str(dsd).strip() if dsd else "—"

        # Normalize whitespace-only cells (Excel sometimes uses \xa0)
        if center_s.replace("\xa0", "").strip() == "":
            center_s = "Unassigned"
        if dsd_s.replace("\xa0", "").strip() == "":
            dsd_s = "—"

        respondents.append({
            "company": company_s,
            "center": center_s,
            "dsd": dsd_s,
            "score": score,
        })
        dist[score - 1] += 1
        center_counts[center_s] += 1

    total = len(respondents)
    if total == 0:
        return None

    promoters = sum(1 for r in respondents if r["score"] >= 9)
    detractors = sum(1 for r in respondents if r["score"] <= 6)
    passives = total - promoters - detractors
    nps_score = round((promoters / total - detractors / total) * 100, 1)
    pro_pct = round(promoters / total * 100, 1)
    det_pct = round(detractors / total * 100, 1)

    # Sort centers by count desc
    by_center = sorted(center_counts.items(), key=lambda kv: (-kv[1], kv[0]))

    return {
        "total": total,
        "nps_score": nps_score,
        "pro_pct": pro_pct,
        "det_pct": det_pct,
        "promoters": promoters,
        "passives": passives,
        "detractors": detractors,
        "dist": dist,
        "center_labels": [c for c, _ in by_center],
        "center_counts": [n for _, n in by_center],
        "respondents": respondents,
    }


# ---------------------------------------------------------------------------
# Customer Churn Parser
# ---------------------------------------------------------------------------
def parse_customer_churn(filepath):
    """Parse Customer Churn Excel → dict with all data needed for customerchurn.html."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["churn"]

    rows = []
    for r in range(3, ws.max_row + 1):
        camp = ws.cell(r, 1).value
        center = ws.cell(r, 2).value
        if not center:
            continue

        dt = ws.cell(r, 5).value
        company = ws.cell(r, 9).value or ""
        prod_quality = ws.cell(r, 10).value
        delivery = ws.cell(r, 11).value
        price = ws.cell(r, 12).value
        billing = ws.cell(r, 13).value
        reached_yes = ws.cell(r, 14).value  # 'X' if yes
        reached_no = ws.cell(r, 15).value   # 'X' if no
        support_exp = ws.cell(r, 16).value
        solution_own = ws.cell(r, 17).value      # 'I source and sharpen my own knives'
        solution_exchange = ws.cell(r, 18).value  # 'Another knife exchange provider'
        verbatim = ws.cell(r, 19).value or ""

        rows.append({
            "camp": str(camp).strip() if camp else "",
            "center": str(center).strip(),
            "date": dt,
            "company": str(company).strip(),
            "prod_quality": to_num(prod_quality),
            "delivery": to_num(delivery),
            "price": to_num(price),
            "billing": to_num(billing),
            "reached_out": bool(reached_yes and str(reached_yes).strip()),
            "support_exp": to_num(support_exp),
            "solution_own": bool(solution_own and str(solution_own).strip()),
            "solution_exchange": bool(solution_exchange and str(solution_exchange).strip()),
            "verbatim": str(verbatim).strip(),
        })

    if not rows:
        return None

    total = len(rows)

    # Date range
    dates = []
    for r in rows:
        if r["date"]:
            if hasattr(r["date"], "strftime"):
                dates.append(r["date"])
    date_min = min(dates).strftime("%b %d, %Y") if dates else "N/A"
    date_max = max(dates).strftime("%b %d, %Y") if dates else "N/A"

    # Satisfaction dimensions (means of 1-5 scores)
    dims = {"prod_quality": [], "delivery": [], "price": [], "billing": []}
    for r in rows:
        for k in dims:
            if r[k] is not None and 1 <= r[k] <= 5:
                dims[k].append(r[k])
    dim_means = {k: round(sum(v)/len(v), 2) if v else 0 for k, v in dims.items()}

    # Stacked satisfaction distribution (1-5 per dimension)
    dim_dist = {}
    for k in dims:
        dim_dist[k] = [sum(1 for v in dims[k] if v == i) for i in range(1, 6)]

    # Churn by center
    center_counts = defaultdict(int)
    center_sat = defaultdict(list)
    center_reached = defaultdict(lambda: [0, 0])  # [reached, total]
    for r in rows:
        c = r["center"]
        center_counts[c] += 1
        sat_vals = [r[k] for k in ["prod_quality", "delivery", "price", "billing"] if r[k] is not None]
        if sat_vals:
            center_sat[c].extend(sat_vals)
        center_reached[c][1] += 1
        if r["reached_out"]:
            center_reached[c][0] += 1

    center_sorted = sorted(center_counts.items(), key=lambda x: x[1], reverse=True)
    drill_center = []
    for name, n in center_sorted:
        avg_sat = round(sum(center_sat[name]) / len(center_sat[name]), 2) if center_sat[name] else 0
        reached_pct = round(center_reached[name][0] / center_reached[name][1] * 100) if center_reached[name][1] else 0
        drill_center.append({"center": name, "n": n, "sat": f"{avg_sat:.2f}", "reached": f"{reached_pct}%"})

    # Reached out before canceling
    reached_total = sum(1 for r in rows if r["reached_out"])
    reached_pct = round(reached_total / total * 100)

    # Support experience (only for those who reached out)
    support_scores = [r["support_exp"] for r in rows if r["support_exp"] is not None and 1 <= r["support_exp"] <= 5]
    support_mean = round(sum(support_scores) / len(support_scores), 2) if support_scores else 0
    support_dist = [sum(1 for s in support_scores if s == i) for i in range(1, 6)]

    # Current knife solution
    sol_exchange = sum(1 for r in rows if r["solution_exchange"])
    sol_own = sum(1 for r in rows if r["solution_own"])
    sol_none = total - sol_exchange - sol_own

    # Verbatim feedback
    feedback = []
    for r in rows:
        if r["verbatim"] and len(r["verbatim"]) > 5:
            sentiment = classify_sentiment(r["verbatim"])
            feedback.append({
                "company": r["company"],
                "center": r["center"],
                "feedback": r["verbatim"].replace('"', '\\"').replace("\n", " "),
                "sentiment": sentiment,
            })

    return {
        "total": total,
        "dim_means": dim_means,
        "dim_dist": dim_dist,
        "center_labels": [c[0] for c in center_sorted],
        "center_counts": [c[1] for c in center_sorted],
        "drill_center": drill_center,
        "reached_pct": reached_pct,
        "reached_total": reached_total,
        "support_mean": support_mean,
        "support_n": len(support_scores),
        "support_dist": support_dist,
        "solution": {"exchange": sol_exchange, "own": sol_own, "none": sol_none},
        "feedback": feedback,
        "date_min": date_min,
        "date_max": date_max,
    }


# ---------------------------------------------------------------------------
# Product Churn Parser
# ---------------------------------------------------------------------------
def parse_product_churn(filepath):
    """Parse Product Churn Excel → dict with all data needed for productchurn.html."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["product_churn"]

    # Row 2 has the reason sub-headers (cols W-AC)
    reason_labels = []
    for c in range(23, 30):
        val = ws.cell(2, c).value
        if val:
            reason_labels.append(str(val).strip())

    rows = []
    for r in range(3, ws.max_row + 1):
        center = ws.cell(r, 1).value or ws.cell(r, 21).value
        if not center:
            continue

        dt = ws.cell(r, 4).value
        company = ws.cell(r, 8).value or ""
        product = ws.cell(r, 16).value or ws.cell(r, 15).value or ""
        days_active = ws.cell(r, 20).value
        verbatim = ws.cell(r, 30).value or ""

        # Reasons: check each col (23-29) for 'X'
        reasons = []
        for i, c in enumerate(range(23, 23 + len(reason_labels))):
            val = ws.cell(r, c).value
            if val and str(val).strip().upper() == "X":
                reasons.append(reason_labels[i])

        # Handle "Other" in verbatim col
        other_val = ws.cell(r, 29).value
        if other_val and str(other_val).strip() and str(other_val).strip().upper() != "X":
            if not reasons:
                reasons.append("Other")

        rows.append({
            "center": str(center).strip(),
            "date": dt,
            "company": str(company).strip(),
            "product": str(product).strip(),
            "days_active": int(days_active) if days_active else None,
            "reasons": reasons,
            "verbatim": str(verbatim).strip() if verbatim else "",
        })

    if not rows:
        return None

    total = len(rows)

    # Date range
    dates = []
    for r in rows:
        if r["date"] and hasattr(r["date"], "strftime"):
            dates.append(r["date"])
    date_min = min(dates).strftime("%b %d") if dates else "N/A"
    date_max = max(dates).strftime("%b %d, %Y") if dates else "N/A"

    # Reason counts (priority: standard reasons first, Other only if no standard)
    standard_reasons = ["No longer needed", "Price / budget", "Switched to another vendor",
                        "Product performance", "Never agreed to start the service", "Menu or equipment changes"]
    reason_counts = defaultdict(int)
    for r in rows:
        std = [rr for rr in r["reasons"] if rr in standard_reasons]
        if std:
            for rr in std:
                reason_counts[rr] += 1
        elif r["reasons"]:
            reason_counts["Other (exclusive)"] += 1

    # Cancellations by center
    center_counts = defaultdict(int)
    for r in rows:
        center_counts[r["center"]] += 1
    center_sorted = sorted(center_counts.items(), key=lambda x: x[1], reverse=True)

    # Product breakdown
    product_counts = defaultdict(int)
    for r in rows:
        p = r["product"]
        p_lower = p.lower()
        # Map product codes and names to display categories
        if "slicer" in p_lower or "slcr" in p_lower:
            product_counts["Slicer Blade"] += 1
        elif "processor" in p_lower or p_lower.startswith("fp-"):
            product_counts["Food Processor Blade"] += 1
        elif "can opener" in p_lower or "can-" in p_lower or "sscop" in p_lower:
            product_counts["Can Opener"] += 1
        elif "steel" in p_lower:
            product_counts["Sharpening Steel"] += 1
        elif "grinder" in p_lower:
            product_counts["Grinder Plate"] += 1
        else:
            product_counts[p if p else "Other"] += 1
    product_sorted = sorted(product_counts.items(), key=lambda x: x[1], reverse=True)

    # Tenure distribution
    tenure_buckets = {"0-30 days": 0, "31-90 days": 0, "91-180 days": 0,
                      "181-365 days": 0, "1-2 years": 0, "2+ years": 0}
    days_list = []
    for r in rows:
        d = r["days_active"]
        if d is not None:
            days_list.append(d)
            if d <= 30: tenure_buckets["0-30 days"] += 1
            elif d <= 90: tenure_buckets["31-90 days"] += 1
            elif d <= 180: tenure_buckets["91-180 days"] += 1
            elif d <= 365: tenure_buckets["181-365 days"] += 1
            elif d <= 730: tenure_buckets["1-2 years"] += 1
            else: tenure_buckets["2+ years"] += 1

    avg_days = round(sum(days_list) / len(days_list)) if days_list else 0
    median_days = sorted(days_list)[len(days_list) // 2] if days_list else 0
    early_cancel = sum(1 for d in days_list if d < 90)
    early_pct = round(early_cancel / len(days_list) * 100) if days_list else 0

    # Top reason
    top_reason = max(reason_counts.items(), key=lambda x: x[1]) if reason_counts else ("N/A", 0)

    # Feedback
    feedback = []
    for r in rows:
        if r["verbatim"] and len(r["verbatim"]) > 3 and r["verbatim"].lower() not in ("no", "no.", "n/a", "n/a.", "na"):
            sentiment = classify_sentiment(r["verbatim"])
            feedback.append({
                "company": r["company"],
                "product": r["product"],
                "center": r["center"],
                "feedback": r["verbatim"].replace('"', '\\"').replace("\n", " "),
                "sentiment": sentiment,
            })

    return {
        "total": total,
        "reason_counts": dict(reason_counts),
        "center_labels": [c[0] for c in center_sorted],
        "center_counts": [c[1] for c in center_sorted],
        "product_labels": [p[0] for p in product_sorted],
        "product_counts": [p[1] for p in product_sorted],
        "tenure_buckets": list(tenure_buckets.values()),
        "avg_days": avg_days,
        "median_days": median_days,
        "early_cancel": early_cancel,
        "early_pct": early_pct,
        "top_reason": top_reason[0],
        "top_reason_count": top_reason[1],
        "feedback": feedback,
        "date_min": date_min,
        "date_max": date_max,
        "n_centers": len(center_counts),
    }


# ---------------------------------------------------------------------------
# CES Parser
# ---------------------------------------------------------------------------
def parse_ces_file(filepath, ces_type):
    """Parse a single CES survey file → dict of scores and metadata."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]

    # Row 2 has actual headers, data starts row 3
    scores = []
    centers = []
    dsds = []
    for r in range(3, ws.max_row + 1):
        center = ws.cell(r, 1).value
        score_val = ws.cell(r, 13).value  # CES column is always last (col M = 13)
        dsd = ws.cell(r, 12).value
        if score_val is not None:
            try:
                scores.append(int(score_val))
                if center:
                    centers.append(str(center).strip())
                if dsd:
                    dsds.append(str(dsd).strip())
            except (ValueError, TypeError):
                pass

    if not scores:
        return None

    # Score distribution (1-7)
    dist = [sum(1 for s in scores if s == i) for i in range(1, 8)]

    return {
        "type": ces_type,
        "mean": round(sum(scores) / len(scores), 2),
        "n": len(scores),
        "dist": dist,
        "scores": scores,
        "centers": centers,
        "dsds": dsds,
    }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def to_num(val):
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        try:
            return float(val)
        except (ValueError, TypeError):
            return None


def classify_sentiment(text):
    """Keyword-based sentiment classification for customer feedback."""
    t = text.lower()

    pos_words = ["great", "excellent", "love", "awesome", "top notch", "pleased",
                 "very pleased", "good job", "wonderful", "fantastic", "best",
                 "happy", "thank", "amazing", "perfect", "outstanding"]
    neg_words = ["poor", "terrible", "horrible", "worst", "never again", "steal",
                 "theft", "lied", "lie", "bait and switch", "not sharp", "disappointed",
                 "frustrat", "unacceptable", "ridiculous", "rude", "unprofessional",
                 "overcharg", "rip off", "scam", "don't recommend"]
    action_words = ["could have", "should have", "didn't", "never received",
                    "waiting on", "fix", "follow through", "didn't fit",
                    "never fit", "wrong", "incorrect", "charged for",
                    "getting charged", "discussed", "agreed", "promised",
                    "asked for", "never did", "never got", "multiple occurrences",
                    "not available", "don't offer", "stopped using",
                    "should cancel", "properly equip", "different model",
                    "swapping out", "prior to", "mistranslation",
                    "explain", "communicate", "follow up"]

    pos = sum(1 for w in pos_words if w in t)
    neg = sum(1 for w in neg_words if w in t)
    act = sum(1 for w in action_words if w in t)

    # Closed business / no longer need = neutral (not negative)
    neutral_phrases = ["closed our business", "closed my business", "no longer need",
                       "don't need", "do not need", "budget", "end of the year",
                       "just don't need", "will return", "not that busy"]
    is_neutral = any(p in t for p in neutral_phrases)

    if act >= 1 and not is_neutral:
        return "actionable"
    if neg > pos and not is_neutral:
        return "negative"
    if pos >= 1 and neg == 0:
        return "positive"
    return "neutral"


# ===========================================================================
# HTML UPDATERS — inject parsed data into dashboard HTML
# ===========================================================================

def update_nps_html(data):
    """Update voc/nps.html with parsed NPS data."""
    html_path = VOC_DIR / "nps.html"
    html = html_path.read_text()

    # Update date badge
    html = re.sub(
        r'(Survey Period:)[^<]+',
        f'\\1 {data["date_min"]} &ndash; {data["date_max"]}',
        html
    )

    # NPS classification
    nps = data["nps"]
    if nps < 0: nps_class, nps_label = "alert", "Poor"
    elif nps <= 20: nps_class, nps_label = "neutral", "Good"
    elif nps <= 50: nps_class, nps_label = "positive", "Strong"
    elif nps <= 70: nps_class, nps_label = "positive", "Exceptional"
    else: nps_class, nps_label = "positive", "World Class"

    nps_range_map = {"Poor": "below 0", "Good": "0&ndash;20 range", "Strong": "21&ndash;50 range",
                     "Exceptional": "51&ndash;70 range", "World Class": "above 70"}

    # KPI 1: NPS Score
    html = re.sub(
        r'(<div class="kpi-card )\w+(">\s*<div class="label">NPS Score</div>\s*<div class="value">)[^<]+(</div>\s*<div class="detail">)[^<]+(</div>)',
        f'\\g<1>{nps_class}\\2{"+" if nps > 0 else ""}{nps}\\3{nps_label} ({nps_range_map[nps_label]})\\4',
        html
    )

    # KPI 2: Total Responses
    html = re.sub(
        r'(<div class="label">Total Responses</div>\s*<div class="value">)\d+(</div>\s*<div class="detail">)[^<]+(</div>)',
        f'\\g<1>{data["total"]}\\2Existing customers surveyed\\3',
        html
    )

    # KPI 3: Promoters %
    pro_pct = round(data["promoters"] / data["total"] * 100, 1)
    html = re.sub(
        r'(<div class="label">Promoters[^<]*</div>\s*<div class="value">)[\d.]+%(</div>\s*<div class="detail">)\d+ of \d+( respondents</div>)',
        lambda m: f'{m.group(1)}{pro_pct}%{m.group(2)}{data["promoters"]} of {data["total"]}{m.group(3)}',
        html
    )

    # KPI 4: Detractors %
    det_pct = round(data["detractors"] / data["total"] * 100, 1)
    html = re.sub(
        r'(<div class="label">Detractors[^<]*</div>\s*<div class="value">)[\d.]+%(</div>\s*<div class="detail">)\d+ of \d+( respondents</div>)',
        lambda m: f'{m.group(1)}{det_pct}%{m.group(2)}{data["detractors"]} of {data["total"]}{m.group(3)}',
        html
    )

    # Update score distribution chart data
    html = re.sub(
        r"(datasets:\s*\[\{\s*data:\s*)\[[\d,\s]+\](\s*,\s*backgroundColor:\s*distColors)",
        lambda m: f"{m.group(1)}{data['dist']}{m.group(2)}",
        html
    )

    # Update promoter/passive/detractor doughnut
    html = re.sub(
        r"(labels:\s*\['Promoters \(9-10\)','Passives \(7-8\)','Detractors \(1-6\)'\]\s*,\s*datasets:\s*\[\{\s*data:\s*)\[[\d,\s]+\]",
        lambda m: f"{m.group(1)}[{data['promoters']}, {data['passives']}, {data['detractors']}]",
        html
    )
    # Fix total in tooltip
    html = re.sub(r"ctx\.raw/\d+\*100", lambda m: f"ctx.raw/{data['total']}*100", html)

    # Update centerData
    center_js = json.dumps(data["center_data"])
    center_js = re.sub(r'"(\w+)":', r'\1:', center_js)
    html = re.sub(
        r"const centerData = \[.*?\];",
        lambda m: f"const centerData = {center_js};",
        html, flags=re.DOTALL
    )

    # Update drillCenter
    dc_js = json.dumps(data["drill_center"])
    dc_js = re.sub(r'"(\w+)":', r'\1:', dc_js)
    html = re.sub(
        r"const drillCenter = \[.*?\];",
        lambda m: f"const drillCenter = {dc_js};",
        html, flags=re.DOTALL
    )

    # Update drillDSD
    dd_js = json.dumps(data["drill_dsd"])
    dd_js = re.sub(r'"(\w+)":', r'\1:', dd_js)
    html = re.sub(
        r"const drillDSD = \[.*?\];",
        lambda m: f"const drillDSD = {dd_js};",
        html, flags=re.DOTALL
    )

    # Update drillAM — only if parser provided it (new Excel format has no AM data → preserve existing HTML)
    if data.get("drill_am") is not None:
        da_js = json.dumps(data["drill_am"])
        da_js = re.sub(r'"(\w+)":', r'\1:', da_js)
        html = re.sub(
            r"const drillAM = \[.*?\];",
            lambda m: f"const drillAM = {da_js};",
            html, flags=re.DOTALL
        )
        am_note = ""
    else:
        am_note = " (AM data preserved from HTML)"

    html_path.write_text(html)
    log.info(f"  Updated nps.html — {data['total']} responses, NPS {data['nps']}{am_note}")


def update_nps_new_customers_section(data):
    """Update the 'New Customer NPS' section at the bottom of nps.html.

    Touches only the new-customer stat values, the two new-customer charts, and
    the respondent table. The top/existing-customer portion of nps.html is not
    modified by this function.
    """
    html_path = VOC_DIR / "nps.html"
    html = html_path.read_text()

    nps_str = f"+{data['nps_score']}" if data['nps_score'] >= 0 else f"{data['nps_score']}"
    nps_color = "var(--cozzini-green)" if data['nps_score'] >= 0 else "var(--cozzini-red)"

    # --- KPI stats: replace only the inner contents of <div class="new-cust-row"> ---
    # Anchor the end of the match on the next section (<div class="new-chart-pair">)
    # so we consume the row + card closings but preserve them verbatim.
    stats_inner = (
        "\n"
        f'      <div class="new-cust-stat">\n'
        f'        <div class="val" style="color:{nps_color};">{nps_str}</div>\n'
        f'        <div class="lbl">NPS Score</div>\n'
        f'      </div>\n'
        f'      <div class="new-cust-stat">\n'
        f'        <div class="val">{data["total"]}</div>\n'
        f'        <div class="lbl">Responses</div>\n'
        f'      </div>\n'
        f'      <div class="new-cust-stat">\n'
        f'        <div class="val">{data["pro_pct"]}%</div>\n'
        f'        <div class="lbl">Promoters (9&ndash;10)</div>\n'
        f'      </div>\n'
        f'      <div class="new-cust-stat">\n'
        f'        <div class="val" style="color:var(--cozzini-red);">{data["det_pct"]}%</div>\n'
        f'        <div class="lbl">Detractors (1&ndash;6)</div>\n'
        f'      </div>\n'
        f'    '
    )
    html = re.sub(
        r'(<div class="new-cust-row">)(.*?)(</div>\s*</div>\s*\n\s*<div class="new-chart-pair")',
        lambda m: m.group(1) + stats_inner + m.group(3),
        html, count=1, flags=re.DOTALL
    )

    # --- Score distribution chart (newCustDistChart) ---
    dist_js = json.dumps(data["dist"])
    html = re.sub(
        r"(datasets: \[\{ data: )\[[^\]]*\](, backgroundColor: newCustDistColors)",
        lambda m: f"{m.group(1)}{dist_js}{m.group(2)}",
        html
    )

    # --- By-center chart (newCustCenterChart) ---
    center_labels_js = json.dumps(data["center_labels"])
    center_counts_js = json.dumps(data["center_counts"])
    # Find the newCustCenterChart block and swap labels + data arrays
    def _replace_center_chart(match):
        block = match.group(0)
        block = re.sub(r"labels: \[[^\]]*\]", f"labels: {center_labels_js}", block, count=1)
        block = re.sub(r"(datasets: \[\{ data: )\[[^\]]*\]", lambda m: f"{m.group(1)}{center_counts_js}", block, count=1)
        return block
    html = re.sub(
        r"new Chart\(document\.getElementById\('newCustCenterChart'\),\s*\{.*?\}\);",
        _replace_center_chart,
        html, count=1, flags=re.DOTALL
    )

    # --- Respondent table data (newCustData array) ---
    resp_js = json.dumps(data["respondents"], ensure_ascii=False)
    # Strip quotes around keys to match the existing JS-object style
    resp_js = re.sub(r'"(company|center|dsd|score)":', r'\1:', resp_js)
    html = re.sub(
        r"const newCustData = \[.*?\];",
        lambda m: f"const newCustData = {resp_js};",
        html, count=1, flags=re.DOTALL
    )

    html_path.write_text(html)
    log.info(f"  Updated nps.html new-customer section — {data['total']} responses, NPS {nps_str}")


def update_customer_churn_html(data):
    """Update voc/customerchurn.html with parsed customer churn data."""
    html_path = VOC_DIR / "customerchurn.html"
    html = html_path.read_text()

    # Date badge
    html = re.sub(
        r'(Survey Period:)[^<]+',
        f'\\1 {data["date_min"]} &ndash; {data["date_max"]}',
        html
    )

    # Count unique campaign waves
    camps = set()
    for r in [row for row in [] ]:  # We don't have camp data here — use center count
        pass
    n_centers = len(data["center_labels"])

    # KPI 1: Churned Customers
    html = re.sub(
        r'(<div class="label">Churned Customers</div>\s*<div class="value">)\d+(</div>\s*<div class="detail">)[^<]+(</div>)',
        f'\\g<1>{data["total"]}\\2Across {n_centers} sharpening centers\\3',
        html
    )
    # KPI 2: Avg Satisfaction
    avg_sat = sum(data["dim_means"].values()) / len(data["dim_means"])
    html = re.sub(
        r'(<div class="label">Avg. Satisfaction[^<]*</div>\s*<div class="value">)[\d.]+(</div>)',
        f'\\g<1>{avg_sat:.2f}\\2',
        html
    )
    # KPI 3: Reached Out
    html = re.sub(
        r'(<div class="label">Reached Out[^<]*</div>\s*<div class="value">)\d+%(</div>\s*<div class="detail">)[^<]+(</div>)',
        lambda m: f'{m.group(1)}{data["reached_pct"]}%{m.group(2)}{data["reached_total"]} of {data["total"]} contacted Cozzini first{m.group(3)}',
        html
    )
    # KPI 4: Support Experience
    html = re.sub(
        r'(<div class="label">Support Experience[^<]*</div>\s*<div class="value">)[\d.]+(</div>\s*<div class="detail">)[^<]+(</div>)',
        f'\\g<1>{data["support_mean"]}\\2Mean score of {data["support_n"]} who rated support\\3',
        html
    )

    # Satisfaction by Dimension chart
    dim_data = [data["dim_means"]["delivery"], data["dim_means"]["billing"],
                data["dim_means"]["prod_quality"], data["dim_means"]["price"]]
    html = re.sub(
        r"(labels:\s*\['Delivery\\nExperience'.*?datasets:\s*\[\{\s*data:\s*)\[[\d.,\s]+\]",
        lambda m: f"{m.group(1)}{dim_data}",
        html
    )

    # Churn by center chart
    center_labels_js = json.dumps([c.replace("Hopewell Junction, NY", "Hopewell Jct, NY") for c in data["center_labels"]])
    html = re.sub(
        r"(getElementById\('centerChart'\).*?labels:\s*)\[[^\]]+\](\s*,\s*datasets:\s*\[\{\s*data:\s*)\[[\d,\s]+\]",
        lambda m: f"{m.group(1)}{center_labels_js}{m.group(2)}{data['center_counts']}",
        html,
        flags=re.DOTALL
    )

    # Solution doughnut
    html = re.sub(
        r"(labels:\s*\['Another knife exchange provider','Sharpen own knives','No response'\]\s*,\s*datasets:\s*\[\{\s*data:\s*)\[[\d,\s]+\]",
        lambda m: f"{m.group(1)}[{data['solution']['exchange']}, {data['solution']['own']}, {data['solution']['none']}]",
        html
    )
    # Fix total in tooltip
    html = re.sub(r"ctx\.raw/\d+\*100", lambda m: f"ctx.raw/{data['total']}*100", html)

    # Drilldown tables
    dc_js = json.dumps(data["drill_center"])
    dc_js = re.sub(r'"(\w+)":', r'\1:', dc_js)
    html = re.sub(r"const drillCenter = \[.*?\];", lambda m: f"const drillCenter = {dc_js};", html, flags=re.DOTALL)

    # Support experience chart
    html = re.sub(
        r"(datasets:\s*\[\{\s*data:\s*)\[[\d,\s]+\](\s*,\s*backgroundColor:\s*supportColors)",
        lambda m: f"{m.group(1)}{data['support_dist']}{m.group(2)}",
        html
    )

    # Stacked satisfaction chart
    dd = data["dim_dist"]
    for i, label in enumerate(["prod_quality", "delivery", "price", "billing"]):
        pass  # Would need to rebuild the stacked data — complex regex

    # Feedback table
    fb_js = json.dumps(data["feedback"])
    fb_js = re.sub(r'"(\w+)":', r'\1:', fb_js)
    # Use a function replacement to avoid regex escape issues in feedback text
    html = re.sub(r"const feedback = \[.*?\];", lambda m: f"const feedback = {fb_js};", html, flags=re.DOTALL)

    html_path.write_text(html)
    log.info(f"  Updated customerchurn.html — {data['total']} responses")


def update_product_churn_html(data):
    """Update voc/productchurn.html with parsed product churn data."""
    html_path = VOC_DIR / "productchurn.html"
    html = html_path.read_text()

    # Date badge
    html = re.sub(
        r'(Survey Period:)[^<]+',
        f'\\1 {data["date_min"]} &ndash; {data["date_max"]}',
        html
    )

    # KPI 1: Total Cancellations
    html = re.sub(
        r'(<div class="label">Total Cancellations</div>\s*<div class="value">)\d+(</div>\s*<div class="detail">)[^<]+(</div>)',
        lambda m: f'{m.group(1)}{data["total"]}{m.group(2)}Across {data["n_centers"]} sharpening centers{m.group(3)}',
        html
    )
    # KPI 2: Avg Subscription Duration
    # NOTE: previous regex required " days" suffix in the value div; the
    # template has the unit only in the detail line, so regex silently failed.
    html = re.sub(
        r'(<div class="label">Avg. Product Subscription Duration</div>\s*<div class="value">)\d+(</div>\s*<div class="detail">Median: )\d+',
        lambda m: f'{m.group(1)}{data["avg_days"]}{m.group(2)}{data["median_days"]}',
        html
    )
    # KPI 3: Early Cancellations
    # NOTE: label contains "<br>(&lt;90 days)" — previous [^<]* stopped at
    # <br> and regex failed. Use .*? (non-greedy up to the label close tag).
    html = re.sub(
        r'(<div class="label">Early Product Cancellations.*?</div>\s*<div class="value">)\d+(</div>\s*<div class="detail">)[^<]+(</div>)',
        lambda m: f'{m.group(1)}{data["early_cancel"]}{m.group(2)}{data["early_pct"]}% of all product cancellations{m.group(3)}',
        html
    )
    # KPI 4: Top Reason
    # NOTE: HTML label is "Top Churn Reason" (not "Top Cancellation Reason")
    # and the value div has inline style (font-size:20px; margin-top:14px).
    # Previous regex mismatched both, so KPI 4 silently failed every run.
    html = re.sub(
        r'(<div class="label">Top Churn Reason</div>\s*<div class="value"[^>]*>)[^<]+(</div>\s*<div class="detail">)[^<]+(</div>)',
        lambda m: f'{m.group(1)}{data["top_reason"]}{m.group(2)}{data["top_reason_count"]} responses &mdash; review verbatim feedback below{m.group(3)}',
        html
    )

    # Reason chart
    reason_order = ["No longer needed", "Never agreed to start the service", "Price / budget",
                    "Product performance", "Menu or equipment changes", "Switched to another vendor"]
    reason_data = [data["reason_counts"].get(r, 0) for r in reason_order]
    html = re.sub(
        r"(getElementById\('reasonChart'\).*?datasets:\s*\[\{\s*data:\s*)\[[\d,\s]+\]",
        lambda m: f"{m.group(1)}{reason_data}",
        html,
        flags=re.DOTALL
    )

    # Center chart
    center_labels_js = json.dumps([c.replace("Hopewell Junction, NY", "Hopewell Jct, NY") for c in data["center_labels"]])
    html = re.sub(
        r"(getElementById\('centerChart'\).*?labels:\s*)\[[^\]]+\](\s*,\s*datasets:\s*\[\{\s*data:\s*)\[[\d,\s]+\]",
        lambda m: f"{m.group(1)}{center_labels_js}{m.group(2)}{data['center_counts']}",
        html,
        flags=re.DOTALL
    )

    # Product doughnut
    prod_labels_js = json.dumps(data["product_labels"])
    html = re.sub(
        r"(getElementById\('productChart'\).*?labels:\s*)\[[^\]]+\](\s*,\s*datasets:\s*\[\{\s*data:\s*)\[[\d,\s]+\]",
        lambda m: f"{m.group(1)}{prod_labels_js}{m.group(2)}{data['product_counts']}",
        html,
        flags=re.DOTALL
    )
    html = re.sub(r"ctx\.raw/\d+\*100", lambda m: f"ctx.raw/{data['total']}*100", html)

    # Tenure chart
    html = re.sub(
        r"(getElementById\('tenureChart'\).*?datasets:\s*\[\{\s*data:\s*)\[[\d,\s]+\]",
        lambda m: f"{m.group(1)}{data['tenure_buckets']}",
        html,
        flags=re.DOTALL
    )

    # Feedback
    fb_js = json.dumps(data["feedback"])
    fb_js = re.sub(r'"(\w+)":', r'\1:', fb_js)
    # Use a function replacement to avoid regex escape issues in feedback text
    html = re.sub(r"const feedback = \[.*?\];", lambda m: f"const feedback = {fb_js};", html, flags=re.DOTALL)

    html_path.write_text(html)
    log.info(f"  Updated productchurn.html — {data['total']} cancellations")


def update_ces_html(all_ces_data):
    """Update voc/ces.html with all CES survey data."""
    html_path = VOC_DIR / "ces.html"
    html = html_path.read_text()

    type_map = {
        "price": "Pricing Satisfaction",
        "onboarding": "Onboarding Ease",
        "knife": "Knife Sharpness",
        "driver": "Driver Service",
        "invoice_payment": "Invoice Payment",
        "invoice_understanding": "Invoice Understanding",
    }
    chart_map = {
        "price": "chart_price",
        "onboarding": "chart_onboard",
        "knife": "chart_knife",
        "driver": "chart_driver",
        "invoice_payment": "chart_invoice",
        "invoice_understanding": "chart_understand",
    }

    # Rebuild surveys array
    surveys = []
    for ces_type, label in type_map.items():
        if ces_type in all_ces_data:
            d = all_ces_data[ces_type]
            surveys.append({"name": label, "mean": d["mean"], "n": d["n"]})

    surveys.sort(key=lambda x: x["mean"], reverse=True)
    surveys_js = json.dumps(surveys)
    surveys_js = re.sub(r'"(\w+)":', r'\1:', surveys_js)
    html = re.sub(r"const surveys = \[.*?\];", f"const surveys = {surveys_js};", html, flags=re.DOTALL)

    # Update score distributions
    for ces_type, canvas_id in chart_map.items():
        if ces_type in all_ces_data:
            dist = all_ces_data[ces_type]["dist"]
            html = re.sub(
                rf"buildDistChart\('{canvas_id}',\s*\[[\d,\s]+\]\)",
                f"buildDistChart('{canvas_id}',    {dist})",
                html
            )

    # Rebuild center averages (aggregate across all CES surveys)
    center_scores = defaultdict(list)
    dsd_scores = defaultdict(list)
    for ces_type, d in all_ces_data.items():
        for i, score in enumerate(d["scores"]):
            if i < len(d["centers"]):
                center_scores[d["centers"][i]].append(score)
            if i < len(d["dsds"]):
                dsd_scores[d["dsds"][i]].append(score)

    centers = [{"name": k, "avg": round(sum(v)/len(v), 2), "n": len(v)}
               for k, v in center_scores.items()]
    centers.sort(key=lambda x: x["avg"], reverse=True)
    centers_js = json.dumps(centers)
    centers_js = re.sub(r'"(\w+)":', r'\1:', centers_js)
    html = re.sub(r"const centers = \[.*?\];", f"const centers = {centers_js};", html, flags=re.DOTALL)

    dsds = [{"name": k, "avg": round(sum(v)/len(v), 2), "n": len(v)}
            for k, v in dsd_scores.items()]
    dsds.sort(key=lambda x: x["avg"], reverse=True)
    dsds_js = json.dumps(dsds)
    dsds_js = re.sub(r'"(\w+)":', r'\1:', dsds_js)
    html = re.sub(r"const dsds = \[.*?\];", f"const dsds = {dsds_js};", html, flags=re.DOTALL)

    html_path.write_text(html)
    log.info(f"  Updated ces.html — {len(all_ces_data)} survey types, {sum(d['n'] for d in all_ces_data.values())} total responses")


# ===========================================================================
# Git operations
# ===========================================================================

def git_push(files, message):
    """Stage, commit, and push specific files."""
    try:
        for f in files:
            subprocess.run(["git", "add", f], cwd=DASH_DIR, check=True, capture_output=True)

        result = subprocess.run(
            ["git", "diff", "--cached", "--quiet"],
            cwd=DASH_DIR, capture_output=True
        )
        if result.returncode == 0:
            log.info("  No changes to commit")
            return False

        subprocess.run(
            ["git", "commit", "-m", message],
            cwd=DASH_DIR, check=True, capture_output=True
        )
        subprocess.run(
            ["git", "push", "origin", "main"],
            cwd=DASH_DIR, check=True, capture_output=True
        )
        log.info(f"  Pushed to GitHub: {message}")
        return True
    except subprocess.CalledProcessError as e:
        log.error(f"  Git error: {e.stderr.decode() if e.stderr else e}")
        return False


# ===========================================================================
# File processing orchestrator
# ===========================================================================

def process_file(filepath):
    """Process a single new Excel file."""
    filename = os.path.basename(filepath)
    dash_type, ces_sub = classify_file(filename)

    if dash_type is None:
        log.info(f"  Skipping unrecognized file: {filename}")
        return

    log.info(f"Processing: {filename} → {dash_type}" + (f" ({ces_sub})" if ces_sub else ""))

    updated_files = []

    try:
        if dash_type == "nps":
            data = parse_nps(filepath)
            if data:
                update_nps_html(data)
                updated_files.append("nps.html")

        elif dash_type == "nps_new":
            data = parse_nps_new_customers(filepath)
            if data:
                update_nps_new_customers_section(data)
                updated_files.append("nps.html")

        elif dash_type == "customer_churn":
            data = parse_customer_churn(filepath)
            if data:
                update_customer_churn_html(data)
                updated_files.append("customerchurn.html")

        elif dash_type == "product_churn":
            data = parse_product_churn(filepath)
            if data:
                update_product_churn_html(data)
                updated_files.append("productchurn.html")

        elif dash_type == "ces":
            # CES needs all files together — load existing + new
            ces_data = load_all_ces_data()
            new_data = parse_ces_file(filepath, ces_sub)
            if new_data:
                ces_data[ces_sub] = new_data
                update_ces_html(ces_data)
                updated_files.append("ces.html")

        elif dash_type == "am_reference":
            # AM reference file — caches DSD→AM lookup; re-run NPS update if
            # NPS source file exists so the dashboard picks up fresh AM data
            parse_am_reference(filepath)
            nps_src = DASH_DIR / "nps_existing_customers.xlsx"
            if nps_src.exists():
                log.info("  Re-running NPS parse to apply new AM mapping")
                nps_data = parse_nps(str(nps_src))
                if nps_data:
                    update_nps_html(nps_data)
                    updated_files.append("nps.html")
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
            processed = load_processed()
            processed[filename] = {
                "processed_at": timestamp,
                "dashboard": "am_reference",
                "files_updated": updated_files,
                "signature": _file_signature(filepath),
            }
            save_processed(processed)
            if updated_files:
                git_push(updated_files, f"Auto-update {', '.join(updated_files)} from AM reference refresh ({timestamp})")
            return

        if updated_files:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
            git_push(updated_files, f"Auto-update {', '.join(updated_files)} from {filename} ({timestamp})")

            # Log as processed with a content signature so re-saved files re-trigger
            processed = load_processed()
            processed[filename] = {
                "processed_at": timestamp,
                "dashboard": dash_type,
                "files_updated": updated_files,
                "signature": _file_signature(filepath),
            }
            save_processed(processed)

    except Exception as e:
        log.error(f"  Error processing {filename}: {e}", exc_info=True)


def load_all_ces_data():
    """Load all existing CES Excel files to aggregate."""
    ces_data = {}
    for f in DASH_DIR.glob("*ces*.xlsx"):
        _, sub = classify_file(f.name)
        if sub and sub != "driver_last":
            data = parse_ces_file(str(f), sub)
            if data:
                ces_data[sub] = data
    return ces_data


# ===========================================================================
# File watcher
# ===========================================================================

class DashboardHandler(FileSystemEventHandler):
    """Watch for new Excel files and process them."""

    def __init__(self):
        self.debounce = {}  # filepath → timestamp to avoid double-processing

    def on_created(self, event):
        if event.is_directory:
            return
        self._handle(event.src_path)

    def on_modified(self, event):
        if event.is_directory:
            return
        self._handle(event.src_path)

    def _handle(self, filepath):
        filename = os.path.basename(filepath)
        if not filename.endswith((".xlsx", ".xls")):
            return
        if filename.startswith("~$"):  # Excel temp files
            return

        # Debounce: ignore if we processed this file in the last 30 seconds
        now = time.time()
        if filepath in self.debounce and now - self.debounce[filepath] < 30:
            return
        self.debounce[filepath] = now

        # Wait for file to finish writing (OneDrive sync can be slow)
        time.sleep(3)

        # Skip only if the EXACT same bytes were already processed.
        # Use size+mtime signature so re-saved / replaced files trigger refresh.
        if _already_processed(filepath, filename):
            log.info(f"  Already processed (same content): {filename} — skipping")
            return

        process_file(filepath)


def _file_signature(filepath):
    """Return a stable signature for (size, mtime) so we detect file changes.

    Uses full-precision mtime — truncating to integer seconds caused re-saves
    with identical byte size (common for XLSX after removing a blank/minor row)
    to be misclassified as 'already processed'.
    """
    try:
        st = os.stat(filepath)
        return f"{st.st_size}:{st.st_mtime}"
    except OSError:
        return None


def _already_processed(filepath, filename):
    """True only if the file is logged AND its current signature matches."""
    processed = load_processed()
    entry = processed.get(filename)
    if not entry:
        return False
    # Older entries (no signature field) → re-process once to attach a signature
    prev_sig = entry.get("signature")
    if not prev_sig:
        return False
    return _file_signature(filepath) == prev_sig


def main():
    if "--once" in sys.argv:
        # Process all unprocessed Excel files and exit
        log.info("=== Running once — processing all pending files ===")
        for f in sorted(DASH_DIR.glob("*.xlsx")):
            if not _already_processed(str(f), f.name):
                process_file(str(f))
            else:
                log.info(f"  Already processed (same content): {f.name} — skipping")
        log.info("=== Done ===")
        return

    log.info("=" * 60)
    log.info("Cozzini Dashboard Watcher started")
    log.info(f"Watching: {DASH_DIR}")
    log.info("Drop Excel files here to auto-update dashboards")
    log.info("=" * 60)

    handler = DashboardHandler()
    observer = Observer()
    observer.schedule(handler, str(DASH_DIR), recursive=False)
    observer.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        log.info("Watcher stopped")
        observer.stop()
    observer.join()


if __name__ == "__main__":
    main()
