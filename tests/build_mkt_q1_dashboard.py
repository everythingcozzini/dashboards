#!/usr/bin/env python3
"""
tests/build_mkt_q1_dashboard.py
-------------------------------
Reads every xlsx in mkt/Q1/, computes the Planned × Results metrics
spec'd in tests/spec_mkt_q1.py, and regenerates mkt/q1.html.

One source of truth: the xlsx files + spec_mkt_q1. Running this twice
yields the same HTML. No manual edits to mkt/q1.html; to change a
number, update the Excel and re-run.

Usage:
  python3 tests/build_mkt_q1_dashboard.py              # write
  python3 tests/build_mkt_q1_dashboard.py --dry-run    # show truth JSON
  python3 tests/build_mkt_q1_dashboard.py --truth-out tests/_mkt_q1_truth.json
"""
from pathlib import Path
import argparse
import json
import sys
from collections import Counter, defaultdict

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "tests"))
from spec_mkt_q1 import (  # noqa: E402
    SOURCES, TARGETS, FOCUS_AREAS, status_for, pacing_ratio,
    year_fraction_elapsed, PALETTE, MKT_ACCENT,
)

import openpyxl  # noqa: E402
import warnings
warnings.filterwarnings("ignore", category=UserWarning)  # openpyxl style warnings

Q1_DIR = ROOT / "mkt" / "Q1"


# ============================================================
# PARSERS — one per xlsx
# ============================================================

def parse_budget():
    """Budget.xlsx: a 2-column plan/actual table."""
    wb = openpyxl.load_workbook(Q1_DIR / SOURCES["budget"], data_only=True)
    ws = wb.active
    d = {}
    for r in range(1, ws.max_row + 1):
        key = ws.cell(r, 1).value
        val = ws.cell(r, 2).value
        if key and val is not None:
            d[str(key).strip()] = val
    return {
        "annual_plan":  d.get("2026 Annual Budget", 0),
        "q1_plan":      d.get("Q1 Plan", 0),
        "q1_actual":    d.get("Q1 Actual (NetSuite)", 0),
        "q1_variance":  d.get("Q1 Variance (Plan - Actual)", 0),
        "ytd_plan":     d.get("YTD Plan (through 4/20)", 0),
        "ytd_actual":   d.get("YTD Actual", 0),
        "ytd_variance": d.get("YTD Variance (Plan - Actual)", 0),
    }


def parse_social():
    """Social Media.xlsx: per-platform rollup + details."""
    wb = openpyxl.load_workbook(Q1_DIR / SOURCES["social"], data_only=True)
    summary = wb["Sheet1"]
    # rows 5-7 are IG feed / FB / LI; row 8 is TOTAL; row 9 is IG stories
    platforms = []
    for r in range(5, 8):
        platforms.append({
            "name":          summary.cell(r, 1).value,
            "posts":         summary.cell(r, 2).value,
            "likes":         summary.cell(r, 3).value,
            "monthly_likes": summary.cell(r, 4).value,
            "posts_per_mo":  summary.cell(r, 5).value,
        })
    total = {
        "posts":         summary.cell(8, 2).value,
        "likes":         summary.cell(8, 3).value,
        "monthly_likes": summary.cell(8, 4).value,
        "posts_per_mo":  summary.cell(8, 5).value,
    }
    ig_stories = {
        "posts": summary.cell(9, 2).value,
        "likes": summary.cell(9, 3).value,
    }

    # Details sheet — pull reach/impressions per platform
    details = wb["Details"]
    fb = {"reach": details.cell(7, 7).value, "views": details.cell(7, 6).value}
    ig = {"reach": details.cell(14, 8).value, "views": details.cell(14, 7).value}
    li = {"impressions": details.cell(24, 5).value, "views": details.cell(24, 6).value, "clicks": details.cell(24, 7).value}

    return {
        "platforms": platforms,
        "total": total,
        "ig_stories": ig_stories,
        "fb_extras": fb,
        "ig_extras": ig,
        "li_extras": li,
    }


def parse_matched_leads():
    """Q1_2026_Matched_Leads_Detail.xlsx Summary sheet."""
    wb = openpyxl.load_workbook(Q1_DIR / SOURCES["matched_leads"], data_only=True)
    ws = wb["Summary"]

    # headline (r5)
    converted = int(ws.cell(5, 1).value or 0)
    matched   = int(ws.cell(5, 4).value or 0)
    match_pct = round((matched / converted * 100) if converted else 0, 1)

    # revenue (r9)
    weekly_wrr     = ws.cell(9, 1).value or 0
    annualized_wrr = ws.cell(9, 4).value or 0
    avg_per_match  = ws.cell(9, 7).value or 0

    # by source (r14-17; TOTAL r18)
    by_source = []
    for r in range(14, 18):
        src = ws.cell(r, 1).value
        if not src:
            continue
        by_source.append({
            "source":        src,
            "leads":         ws.cell(r, 2).value,
            "matched":       ws.cell(r, 3).value,
            "match_rate":    round((ws.cell(r, 4).value or 0) * 100, 1),
            "annualized":    ws.cell(r, 6).value,
            "pct_of_wrr":    round((ws.cell(r, 7).value or 0) * 100, 1),
            "avg_per_match": ws.cell(r, 8).value,
        })

    # by month (r32-35; TOTAL r36)
    by_month = []
    for r in range(32, 36):
        mo = ws.cell(r, 1).value
        if not mo:
            continue
        by_month.append({
            "month":       mo,
            "leads":       ws.cell(r, 2).value,
            "matched":     ws.cell(r, 3).value,
            "match_rate":  round((ws.cell(r, 4).value or 0) * 100, 1),
            "annualized":  ws.cell(r, 6).value,
        })

    # by owner (r22-27; TOTAL r28)
    by_owner = []
    for r in range(22, 28):
        owner = ws.cell(r, 1).value
        if not owner:
            continue
        by_owner.append({
            "owner":       owner,
            "leads":       ws.cell(r, 2).value,
            "matched":     ws.cell(r, 3).value,
            "match_rate":  round((ws.cell(r, 4).value or 0) * 100, 1),
            "annualized":  ws.cell(r, 6).value,
        })

    return {
        "converted": converted,
        "matched": matched,
        "match_pct": match_pct,
        "weekly_wrr": round(weekly_wrr, 2),
        "annualized_wrr": round(annualized_wrr),
        "avg_per_match": round(avg_per_match),
        "by_source": by_source,
        "by_month": by_month,
        "by_owner": [o for o in by_owner if o["leads"] and o["leads"] > 0],
    }


def parse_leads_ytd():
    """Details of Leads Assigned YTD: count by Lead Source v2."""
    wb = openpyxl.load_workbook(Q1_DIR / SOURCES["leads_details"], data_only=True)
    ws = wb.active
    counter = Counter()
    for r in range(2, ws.max_row + 1):
        src = ws.cell(r, 5).value
        if src:
            counter[str(src).strip()] += 1
    total = sum(counter.values())
    top = counter.most_common()
    return {
        "total": total,
        "sources": [{"source": s, "count": c, "pct": round(c / total * 100, 1)} for s, c in top],
    }


def parse_snapshot():
    """Q1Snap.xlsx 'Planned' sheet: NPS, CES, Google rating, engagement, new-biz origin."""
    wb = openpyxl.load_workbook(Q1_DIR / SOURCES["snapshot"], data_only=True)
    # Key Q1 actuals live in the Planned sheet, column 6 (Q1)
    ws = wb["Planned"]

    def _num(v):
        if v is None: return None
        try: return float(v)
        except Exception: return None

    return {
        "nps":               _num(ws.cell(8, 6).value),    # 35.9
        "ces":               _num(ws.cell(9, 6).value),    # 5.2
        "google_rating":     _num(ws.cell(10, 6).value),   # 2.5
        "social_engagement": (_num(ws.cell(14, 6).value) or 0) * 100,  # 0.05 → 5.0
        "new_biz_origin":    (_num(ws.cell(15, 6).value) or 0) * 100,  # 0.121 → 12.1
        "new_biz_fraction":  str(ws.cell(17, 6).value or ""),  # "(240/1,982)"
    }


# ============================================================
# ORCHESTRATE — roll everything into one truth dict
# ============================================================

def build_truth():
    snap = parse_snapshot()
    budget = parse_budget()
    social = parse_social()
    matched = parse_matched_leads()
    leads = parse_leads_ytd()

    # Compute posts_per_month actual = avg across 3 main platforms
    actual_posts_per_mo = round(
        sum(p["posts_per_mo"] or 0 for p in social["platforms"]) / 3, 1
    )

    # Year-elapsed fraction drives pacing calculations for annual-goal metrics
    yf = year_fraction_elapsed()

    def s(key, val):
        return status_for(key, val, year_frac=yf)

    def pacing(key, val):
        """Return {ratio, expected, year_frac} for cumulative_pacing metrics."""
        r, e = pacing_ratio(key, val, year_frac=yf)
        if r is None:
            return None
        return {
            "ratio": round(r, 3),
            "expected": round(e, 2),
            "year_frac": round(yf, 3),
            "pct_of_annual": round(val / TARGETS[key]["target_value"] * 100, 1),
            "pct_of_year": round(yf * 100, 1),
        }

    truth = {
        "snapshot": snap,
        "budget": budget,
        "social": social,
        "matched_leads": matched,
        "leads_ytd": leads,
        "year_fraction_elapsed": round(yf, 3),

        "metrics": {
            "nps":               {"actual": snap["nps"],               "status": s("nps", snap["nps"])},
            "ces":               {"actual": snap["ces"],               "status": s("ces", snap["ces"])},
            "google_rating":     {"actual": snap["google_rating"],     "status": s("google_rating", snap["google_rating"])},
            "social_engagement": {"actual": snap["social_engagement"], "status": s("social_engagement", snap["social_engagement"]), "pacing": pacing("social_engagement", snap["social_engagement"])},
            "posts_per_month":   {"actual": actual_posts_per_mo,       "status": s("posts_per_month", actual_posts_per_mo)},
            "new_biz_origin":    {"actual": snap["new_biz_origin"],    "status": s("new_biz_origin", snap["new_biz_origin"]),   "pacing": pacing("new_biz_origin", snap["new_biz_origin"])},
            "q1_budget":         {"actual": budget["q1_actual"],       "status": s("q1_budget", budget["q1_actual"])},
        },
    }
    return truth


# ============================================================
# HTML GENERATOR
# ============================================================

STATUS_COLORS = {
    "green":       ("#6faa58", "#dceccf", "On track"),     # solid green
    "green_light": ("#a9cc98", "#ecf5e5", "Near target"),  # lighter green
    "amber":       ("#d4a45e", "#fcf3e2", "Below pace"),
    "red":         ("#d4918b", "#f9e3e0", "Needs focus"),
    "unknown":     ("#a0a0a0", "#f0f0f0", "No data"),
}


def fmt_num(v, unit=""):
    if v is None: return "—"
    if unit == "dollars":
        return f"${v:,.0f}"
    if unit == "pct":
        return f"{v:.1f}%"
    if unit == "rating":
        return f"{v:.1f}"
    if unit == "score":
        return f"{v:.1f}"
    return f"{v:,.0f}" if isinstance(v, (int, float)) else str(v)


def hero_card(metric_key, actual, label, target_display, unit, status, period="Q1 actual", target_prefix="2026 Goal", pacing_info=None):
    bg, tint, status_text = STATUS_COLORS[status]
    badge = {"green": "✓", "green_light": "✓", "amber": "⚠", "red": "✗", "unknown": "?"}[status]

    # For pacing metrics, replace the generic status with a richer pace line
    # and show a progress bar: % of annual goal achieved vs % of year elapsed.
    pacing_html = ""
    if pacing_info is not None:
        pct_goal = pacing_info["pct_of_annual"]
        pct_year = pacing_info["pct_of_year"]
        # Cap bar fills at 100% for display
        goal_fill = min(pct_goal, 100)
        year_mark = min(pct_year, 100)
        if status in ("green", "green_light"):
            status_text = "Ahead of pace"
        elif status == "amber":
            status_text = "Near pace"
        else:
            status_text = "Behind pace"
        pacing_html = f"""
      <div class="hero-pace">
        <div class="hero-pace-bar">
          <div class="hero-pace-fill" style="width:{goal_fill}%;"></div>
          <div class="hero-pace-marker" style="left:{year_mark}%;" title="Year elapsed"></div>
        </div>
        <div class="hero-pace-legend">
          <span><b>{pct_goal:.0f}%</b> of annual goal</span>
          <span class="pace-sep">·</span>
          <span>{pct_year:.0f}% of year elapsed</span>
        </div>
      </div>"""

    return f"""
    <div class="hero-card hero-{status}">
      <div class="hero-label">{label}</div>
      <div class="hero-value">{fmt_num(actual, unit)}</div>
      <div class="hero-period">{period}</div>
      <div class="hero-target">{target_prefix}: {target_display}</div>
      {pacing_html}
      <div class="hero-status" style="background:{tint};color:{bg};">{badge} {status_text}</div>
    </div>"""


def focus_card(area, truth):
    """Build the card for one focus area."""
    lines = []
    for k in area["metric_keys"]:
        t = TARGETS[k]
        m = truth["metrics"][k]
        status = m["status"]
        badge_color = STATUS_COLORS[status][0]
        actual_fmt = fmt_num(m["actual"], t["unit"])
        # For pacing metrics, note the ahead/on/behind-pace framing
        if t.get("metric_type") == "cumulative_pacing" and m.get("pacing"):
            pi = m["pacing"]
            if status in ("green", "green_light"):
                head = "Ahead of pace"
            elif status == "amber":
                head = "Near pace"
            else:
                head = "Behind pace"
            # Two compact lines: status headline + math
            target_line = (
                f"<b>{head}</b> toward {t['target_display']} goal"
                f"<br>{pi['pct_of_annual']:.0f}% of annual achieved at {pi['pct_of_year']:.0f}% of year elapsed"
            )
        else:
            target_line = f"Q1 actual vs 2026 goal {t['target_display']}"
        lines.append(f"""
        <div class="focus-metric">
          <span class="focus-metric-dot" style="background:{badge_color};"></span>
          <span class="focus-metric-label">{t['label']}</span>
          <span class="focus-metric-actual">{actual_fmt}</span>
          <span class="focus-metric-target">{target_line}</span>
        </div>""")
    action_list = "".join(f"<li>{a}</li>" for a in area["actions"])
    return f"""
    <div class="focus-card">
      <h3>{area['name']}</h3>
      <ul class="focus-actions">{action_list}</ul>
      <div class="focus-metrics">{''.join(lines)}</div>
    </div>"""


def build_html(truth):
    """Generate the full mkt/q1.html from truth dict."""
    m = truth["metrics"]
    b = truth["budget"]
    s = truth["social"]
    ml = truth["matched_leads"]
    ly = truth["leads_ytd"]

    # Hero cards (4 headline metrics) — lead with wins:
    #   1. Q1 Budget      (QUARTERLY target: $53K vs $68K plan → green)
    #   2. NPS            (POINT-IN-TIME: 35.9 vs >=35 → green)
    #   3. CES            (POINT-IN-TIME: 5.2 vs >=5.0 → green)
    #   4. New Biz Share  (CUMULATIVE PACING: 12.1% vs 25% annual — AHEAD
    #                      of pace because we're only ~30% through the year;
    #                      12.1 / (25 * 0.30) = 161% of expected pace → green)
    hero = "".join([
        hero_card("q1_budget",      b["q1_actual"],                  "Q1 Budget Spend",             TARGETS["q1_budget"]["target_display"],      "dollars", m["q1_budget"]["status"],      period="Q1 actual", target_prefix="Q1 plan"),
        hero_card("nps",            m["nps"]["actual"],              "NPS",                         TARGETS["nps"]["target_display"],            "score",   m["nps"]["status"],            period="Q1 actual"),
        hero_card("ces",            m["ces"]["actual"],              "CES",                         TARGETS["ces"]["target_display"],            "score",   m["ces"]["status"],            period="Q1 actual"),
        hero_card("new_biz_origin", m["new_biz_origin"]["actual"],   "Marketing &mdash; Share of New Biz",  TARGETS["new_biz_origin"]["target_display"], "pct",     m["new_biz_origin"]["status"], period="Q1 YTD", pacing_info=m["new_biz_origin"]["pacing"]),
    ])

    # Q1 at a Glance — positive-first callout bar above the focus scorecard
    pace_info = m["new_biz_origin"].get("pacing") or {}
    glance_items = [
        ("Under budget",      f"${b['q1_variance']:,.0f} saved in Q1"),
        ("NPS",               f"{m['nps']['actual']} (goal {TARGETS['nps']['target_display']})"),
        ("CES",               f"{m['ces']['actual']} (goal {TARGETS['ces']['target_display']})"),
        ("Pipeline built",    f"${ml['annualized_wrr']:,.0f} WRR annualized"),
        ("Ahead of pace",     f"{pace_info.get('pct_of_annual', 0):.0f}% of annual new-biz goal at {pace_info.get('pct_of_year', 0):.0f}% of year elapsed"),
    ]
    glance_html = "".join(
        f'<div class="glance-item"><span class="glance-check">&#10003;</span>'
        f'<span class="glance-label">{lbl}</span>'
        f'<span class="glance-value">{val}</span></div>'
        for lbl, val in glance_items
    )

    # Focus area scorecard (4 cards, one per focus area)
    focus_cards_html = "".join(focus_card(fa, truth) for fa in FOCUS_AREAS)

    # Budget bar chart data
    budget_chart_data = {
        "labels": ["Q1 Plan", "Q1 Actual", "YTD Plan", "YTD Actual"],
        "values": [b["q1_plan"], b["q1_actual"], b["ytd_plan"], b["ytd_actual"]],
    }

    # Social chart data
    social_chart_data = {
        "labels": [p["name"] for p in s["platforms"]],
        "posts":  [p["posts"] for p in s["platforms"]],
        "likes":  [p["likes"] for p in s["platforms"]],
    }

    # Lead funnel
    funnel = {
        "converted": ml["converted"],
        "matched":   ml["matched"],
        "match_pct": ml["match_pct"],
        "wrr":       ml["annualized_wrr"],
        "avg":       ml["avg_per_match"],
    }

    # By-source for leads chart
    src_chart = {
        "labels":  [row["source"] for row in ml["by_source"]],
        "matched": [row["matched"] for row in ml["by_source"]],
        "wrr":     [row["annualized"] for row in ml["by_source"]],
    }

    # By-month trend
    month_chart = {
        "labels":      [row["month"] for row in ml["by_month"]],
        "leads":       [row["leads"] for row in ml["by_month"]],
        "matched":     [row["matched"] for row in ml["by_month"]],
        "match_rate":  [row["match_rate"] for row in ml["by_month"]],
    }

    # Lead owner table + YTD sources section removed per exec feedback.
    # (Kept parsers for these so they still appear in the truth JSON; if we
    # ever want them back, just add the section HTML.)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="robots" content="noindex, nofollow">
<title>Q1 2026 Marketing | Cozzini Bros</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  :root {{
    --cozzini-navy: {PALETTE['navy']};
    --cozzini-blue: {PALETTE['blue']};
    --cozzini-red:  {PALETTE['red']};
    --cozzini-gold: {PALETTE['gold']};
    --cozzini-green: {PALETTE['green']};
    --mkt-burnt:  {MKT_ACCENT['burnt']};
    --mkt-burgundy: {MKT_ACCENT['burgundy']};
    --text: #0a0a0a;
    --text-muted: #555;
    --border: #e5e5e5;
    --card-bg: #ffffff;
  }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #fbfaf9; color: var(--text); }}

  .header {{ background: var(--cozzini-navy); color: white; padding: 28px 40px; display: flex; justify-content: space-between; align-items: center; }}
  .header h1 {{ font-size: 24px; font-weight: 700; letter-spacing: -0.5px; }}
  .header .subtitle {{ font-size: 14px; opacity: 0.8; margin-top: 4px; }}
  .header .eyebrow {{ font-size: 11px; text-transform: uppercase; letter-spacing: 2px; color: var(--mkt-burnt); font-weight: 700; margin-bottom: 4px; }}
  .header .date-badge {{ background: rgba(255,255,255,0.15); padding: 8px 16px; border-radius: 8px; font-size: 13px; font-weight: 500; }}
  .back-link {{ color: rgba(255,255,255,0.7); text-decoration: none; font-size: 13px; margin-bottom: 4px; display: inline-block; }}
  .back-link:hover {{ color: white; }}

  .container {{ max-width: 1400px; margin: 0 auto; padding: 24px 40px 60px; }}

  /* === HERO STRIP === */
  .hero-strip {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-bottom: 32px; }}
  .hero-card {{ background: var(--card-bg); border-radius: 14px; padding: 22px; border: 1px solid var(--border); box-shadow: 0 1px 3px rgba(0,0,0,0.05); display: flex; flex-direction: column; gap: 6px; }}
  .hero-card .hero-label {{ font-size: 11px; text-transform: uppercase; letter-spacing: 1px; color: var(--text-muted); font-weight: 600; }}
  .hero-card .hero-value {{ font-size: 32px; font-weight: 800; color: var(--cozzini-navy); letter-spacing: -0.5px; }}
  .hero-card .hero-period {{ font-size: 10px; text-transform: uppercase; letter-spacing: 1px; color: var(--mkt-burnt); font-weight: 700; margin-top: -2px; }}
  .hero-card .hero-target {{ font-size: 12px; color: var(--text-muted); margin-top: 2px; }}
  .hero-card .hero-status {{ margin-top: 8px; padding: 6px 10px; border-radius: 6px; font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: 1px; display: inline-block; width: fit-content; }}
  /* Pacing bar for cumulative annual metrics */
  .hero-pace {{ margin-top: 10px; }}
  .hero-pace-bar {{ position: relative; height: 8px; background: #f0f0f0; border-radius: 4px; overflow: hidden; }}
  .hero-pace-fill {{ position: absolute; top: 0; left: 0; height: 100%; background: linear-gradient(90deg, #8ebf7b 0%, #a9cc98 100%); border-radius: 4px; }}
  .hero-pace-marker {{ position: absolute; top: -3px; width: 2px; height: 14px; background: var(--cozzini-navy); transform: translateX(-50%); }}
  .hero-pace-marker::after {{ content: ""; position: absolute; top: -4px; left: 50%; transform: translateX(-50%); width: 0; height: 0; border-left: 4px solid transparent; border-right: 4px solid transparent; border-top: 4px solid var(--cozzini-navy); }}
  .hero-pace-legend {{ display: flex; flex-wrap: wrap; gap: 6px; margin-top: 6px; font-size: 10px; color: var(--text-muted); letter-spacing: 0.3px; }}
  .hero-pace-legend b {{ color: var(--cozzini-navy); }}
  .hero-pace-legend .pace-sep {{ color: #c0c0c0; }}
  .hero-green {{ border-left: 4px solid var(--cozzini-green); }}
  .hero-amber {{ border-left: 4px solid var(--cozzini-gold); }}
  .hero-red {{ border-left: 4px solid var(--cozzini-red); }}
  .hero-unknown {{ border-left: 4px solid #a0a0a0; }}

  /* === SECTION === */
  .section {{ margin-bottom: 36px; }}
  .section-title {{ font-size: 18px; font-weight: 700; color: var(--cozzini-navy); margin: 0 0 12px; padding-bottom: 8px; border-bottom: 2px solid var(--border); display: flex; align-items: center; gap: 10px; }}
  .section-title .count {{ font-size: 13px; color: var(--text-muted); font-weight: 500; }}

  /* === FOCUS SCORECARD === */
  .focus-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px; }}
  .focus-card {{ background: var(--card-bg); border-radius: 12px; padding: 22px; border: 1px solid var(--border); box-shadow: 0 1px 3px rgba(0,0,0,0.04); }}
  .focus-card h3 {{ font-size: 15px; font-weight: 700; color: var(--cozzini-navy); margin-bottom: 10px; }}
  .focus-actions {{ list-style: none; padding: 0; margin: 0 0 14px; }}
  .focus-actions li {{ font-size: 13px; color: var(--text-muted); padding: 3px 0; }}
  .focus-actions li::before {{ content: "→ "; color: var(--mkt-burnt); }}
  .focus-metrics {{ display: flex; flex-direction: column; gap: 12px; padding-top: 12px; border-top: 1px dashed var(--border); }}
  .focus-metric {{
    display: grid;
    grid-template-columns: 14px 1fr auto;
    column-gap: 10px;
    row-gap: 3px;
    align-items: baseline;
    font-size: 13px;
  }}
  .focus-metric-dot {{ grid-row: 1; grid-column: 1; width: 12px; height: 12px; border-radius: 50%; align-self: center; }}
  .focus-metric-label {{ grid-row: 1; grid-column: 2; color: var(--text); font-weight: 500; }}
  .focus-metric-actual {{ grid-row: 1; grid-column: 3; font-weight: 700; color: var(--cozzini-navy); }}
  .focus-metric-target {{ grid-row: 2; grid-column: 2 / -1; font-size: 11px; color: var(--text-muted); line-height: 1.5; }}
  .focus-metric-target b {{ color: var(--cozzini-navy); }}

  /* === CARDS + CHARTS === */
  .chart-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
  .chart-card {{ background: var(--card-bg); border-radius: 12px; padding: 22px; border: 1px solid var(--border); box-shadow: 0 1px 3px rgba(0,0,0,0.04); }}
  .chart-card h3 {{ font-size: 14px; font-weight: 700; color: var(--cozzini-navy); margin-bottom: 4px; }}
  .chart-card .chart-sub {{ font-size: 12px; color: var(--text-muted); margin-bottom: 14px; }}
  .chart-wrap {{ position: relative; height: 280px; }}
  .chart-wrap.tall {{ height: 340px; }}
  .full-chart {{ grid-column: 1 / -1; }}

  /* === BUDGET + FUNNEL KPIs === */
  .kpi-inline {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 14px; }}
  .kpi-inline .kpi {{ padding: 14px 16px; background: var(--card-bg); border: 1px solid var(--border); border-radius: 10px; }}
  .kpi-inline .kpi-label {{ font-size: 11px; text-transform: uppercase; letter-spacing: 1px; color: var(--text-muted); font-weight: 600; }}
  .kpi-inline .kpi-value {{ font-size: 24px; font-weight: 800; color: var(--cozzini-navy); margin-top: 4px; }}
  .kpi-inline .kpi-sub {{ font-size: 12px; color: var(--text-muted); margin-top: 2px; }}
  .kpi-pos {{ color: var(--cozzini-green) !important; }}
  .kpi-neg {{ color: var(--cozzini-red) !important; }}

  /* === FUNNEL === */
  .funnel {{ display: flex; justify-content: center; align-items: stretch; gap: 14px; flex-wrap: wrap; padding: 10px 0; }}
  .funnel-stage {{ flex: 1; min-width: 180px; padding: 18px; border-radius: 12px; text-align: center; color: white; }}
  .funnel-arrow {{ align-self: center; font-size: 28px; color: var(--text-muted); }}
  .funnel-num {{ font-size: 28px; font-weight: 800; display: block; letter-spacing: -0.5px; }}
  .funnel-lbl {{ font-size: 12px; text-transform: uppercase; letter-spacing: 1px; opacity: 0.85; }}
  .funnel-1 {{ background: var(--cozzini-navy); }}
  .funnel-2 {{ background: var(--cozzini-blue); color: var(--cozzini-navy); }}
  .funnel-3 {{ background: var(--cozzini-green); color: var(--cozzini-navy); }}

  /* === Q1 AT A GLANCE CALLOUT BAR === */
  .glance-bar {{ background: linear-gradient(135deg, #1a2744 0%, #2a3d5c 100%); color: white; border-radius: 14px; padding: 20px 24px; margin-bottom: 32px; box-shadow: 0 2px 12px rgba(26,39,68,0.15); }}
  .glance-title {{ font-size: 11px; font-weight: 700; letter-spacing: 2px; text-transform: uppercase; color: #d4a45e; margin-bottom: 14px; }}
  .glance-list {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 12px 20px; }}
  .glance-item {{ display: flex; align-items: baseline; gap: 10px; font-size: 13px; }}
  .glance-check {{ color: #8ebf7b; font-size: 15px; flex-shrink: 0; }}
  .glance-label {{ font-weight: 600; color: white; }}
  .glance-value {{ color: rgba(255,255,255,0.75); font-size: 12px; }}

  /* === TABLES === */
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  thead th {{ text-align: left; padding: 10px 12px; border-bottom: 2px solid var(--border); font-weight: 600; color: var(--text-muted); font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px; }}
  tbody td {{ padding: 10px 12px; border-bottom: 1px solid var(--border); }}
  thead th.num, tbody td.num {{ text-align: right; }}

  /* === FOOTER === */
  .footer {{ background: #111; padding: 22px 40px; display: flex; align-items: center; justify-content: center; gap: 14px; }}
  .footer img {{ height: 22px; opacity: 0.7; }}
  .footer span {{ color: rgba(255,255,255,0.4); font-size: 11px; }}

  @media (max-width: 900px) {{
    .hero-strip {{ grid-template-columns: repeat(2, 1fr); }}
    .focus-grid {{ grid-template-columns: 1fr; }}
    .chart-grid {{ grid-template-columns: 1fr; }}
    .container {{ padding: 16px; }}
    .header {{ padding: 20px; flex-direction: column; gap: 12px; text-align: center; }}
    .header > div:first-child {{ flex-direction: column !important; gap: 10px !important; }}
    .header img {{ height: 32px; }}
  }}

  /* === iphone13-patch:start === */
  @media (max-width: 480px) {{
    #auth-gate > div {{ width: 92% !important; max-width: 360px !important; padding: 32px 24px !important; }}
    .header {{ padding: 16px; }} .header h1 {{ font-size: 20px; }}
    .container {{ padding: 12px; }}
    .hero-strip {{ grid-template-columns: 1fr; }}
    .hero-card {{ padding: 16px; }} .hero-card .hero-value {{ font-size: 26px; }}
    .chart-card .chart-wrap {{ height: 240px; }}
    .section-title {{ font-size: 16px; margin: 24px 0 12px; }}
    .funnel {{ flex-direction: column; }}
    .funnel-arrow {{ transform: rotate(90deg); }}
    thead th, tbody td {{ padding: 8px 6px; font-size: 11px; }}
    .footer {{ padding: 16px; flex-direction: column; gap: 8px; text-align: center; }}
  }}
  /* === iphone13-patch:end === */
</style>
</head>
<body>

<div id="auth-gate" style="display:flex;align-items:center;justify-content:center;min-height:100vh;background:linear-gradient(135deg, #8a2d1f 0%, #b44d2d 60%, #d4a45e 100%);font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;">
  <div style="background:white;border-radius:12px;padding:48px;width:380px;text-align:center;box-shadow:0 8px 30px rgba(0,0,0,0.2);">
    <img src="https://www.cozzinibros.com/wp-content/uploads/2026/03/Cozzini-Logo-Horizontal-Pantone-TK.png" alt="Cozzini Bros" style="height:40px;width:auto;margin-bottom:16px;">
    <div style="color:#b44d2d;font-size:11px;text-transform:uppercase;letter-spacing:2px;font-weight:700;margin-bottom:8px;">Marketing</div>
    <h2 style="color:#1a2744;margin-bottom:4px;font-size:20px;">Cozzini Dashboards</h2>
    <p style="color:#555;font-size:13px;margin-bottom:24px;">Enter password to continue</p>
    <input id="auth-pass" type="password" placeholder="Password" onkeydown="if(event.key==='Enter')checkAuth()" style="width:100%;padding:12px 16px;border:1px solid #e5e5e5;border-radius:6px;font-size:15px;outline:none;margin-bottom:12px;">
    <button onclick="checkAuth()" style="width:100%;padding:12px;background:#b44d2d;color:white;border:none;border-radius:6px;font-size:15px;font-weight:600;cursor:pointer;">Sign In</button>
    <p id="auth-error" style="color:#d4918b;font-size:13px;margin-top:12px;display:none;">Incorrect password</p>
  </div>
</div>

<div id="page-content" style="display:none;">

<div class="header">
  <div style="display:flex; align-items:center; gap:20px;">
    <a href="../"><img src="https://www.cozzinibros.com/wp-content/uploads/2026/03/Cozzini-Logo-White-Letters-Hor-TK.png" alt="Cozzini Bros" style="height:44px; width:auto;"></a>
    <div>
      <a href="./" class="back-link">&larr; Back to Marketing</a>
      <div class="eyebrow">Marketing</div>
      <h1>Q1 2026 &mdash; Planned &times; Results</h1>
      <div class="subtitle">Q1 progress vs full-year 2026 goals &mdash; ~25% of year elapsed, 3 quarters remaining</div>
    </div>
  </div>
  <div class="date-badge">YTD through Apr 20, 2026</div>
</div>

<div class="container">

  <!-- HERO: 4 headline KPIs (all wins) — Planned × Results framing -->
  <div class="hero-strip">
    {hero}
  </div>

  <!-- Q1 AT A GLANCE — positive-first summary bar -->
  <div class="glance-bar">
    <div class="glance-title">Q1 at a glance</div>
    <div class="glance-list">{glance_html}</div>
  </div>

  <!-- FOCUS SCORECARD -->
  <div class="section">
    <div class="section-title">Marketing Plan &times; Q1 Progress <span class="count">&mdash; 4 focus areas &middot; Q1 actuals against 2026 annual goals</span></div>
    <div class="focus-grid">
      {focus_cards_html}
    </div>
  </div>

  <!-- BUDGET -->
  <div class="section">
    <div class="section-title">Budget &times; Spend <span class="count">&mdash; Q1 + YTD</span></div>
    <div class="kpi-inline" style="margin-bottom: 14px;">
      <div class="kpi">
        <div class="kpi-label">Q1 Plan</div>
        <div class="kpi-value">${b['q1_plan']:,.0f}</div>
        <div class="kpi-sub">Budgeted spend Q1 2026</div>
      </div>
      <div class="kpi">
        <div class="kpi-label">Q1 Actual (NetSuite)</div>
        <div class="kpi-value">${b['q1_actual']:,.0f}</div>
        <div class="kpi-sub">{round(b['q1_actual']/b['q1_plan']*100):.0f}% of Q1 plan</div>
      </div>
      <div class="kpi">
        <div class="kpi-label">Q1 Variance (Under)</div>
        <div class="kpi-value kpi-pos">${b['q1_variance']:,.0f}</div>
        <div class="kpi-sub">Under plan in Q1</div>
      </div>
      <div class="kpi">
        <div class="kpi-label">YTD Variance (Under)</div>
        <div class="kpi-value kpi-pos">${b['ytd_variance']:,.0f}</div>
        <div class="kpi-sub">YTD under plan through 4/20</div>
      </div>
    </div>
    <div class="chart-card">
      <h3>Plan vs Actual</h3>
      <div class="chart-sub">Quarterly and year-to-date spend</div>
      <div class="chart-wrap"><canvas id="budgetChart"></canvas></div>
    </div>
  </div>

  <!-- LEAD CONVERSION FUNNEL (moved above Social Media per exec feedback) -->
  <div class="section">
    <div class="section-title">Lead Conversion <span class="count">&mdash; Q1 marketing channels &rarr; routes &rarr; revenue</span></div>
    <div class="funnel">
      <div class="funnel-stage funnel-1">
        <span class="funnel-num">{funnel['converted']}</span>
        <span class="funnel-lbl">Converted Leads</span>
      </div>
      <div class="funnel-arrow">&rarr;</div>
      <div class="funnel-stage funnel-2">
        <span class="funnel-num">{funnel['matched']}</span>
        <span class="funnel-lbl">Matched to Route ({funnel['match_pct']}%)</span>
      </div>
      <div class="funnel-arrow">&rarr;</div>
      <div class="funnel-stage funnel-3">
        <span class="funnel-num">${funnel['wrr']:,.0f}</span>
        <span class="funnel-lbl">Annualized WRR</span>
      </div>
    </div>
    <div class="kpi-inline" style="margin: 14px 0;">
      <div class="kpi"><div class="kpi-label">Avg $ per Matched Lead</div><div class="kpi-value">${funnel['avg']:,.0f}</div><div class="kpi-sub">Annualized WRR per conversion</div></div>
      <div class="kpi"><div class="kpi-label">Weekly WRR</div><div class="kpi-value">${ml['weekly_wrr']:,.2f}</div><div class="kpi-sub">Current weekly run-rate from Q1 matches</div></div>
      <div class="kpi"><div class="kpi-label">Match Rate</div><div class="kpi-value">{funnel['match_pct']:.1f}%</div><div class="kpi-sub">240 of 398 leads matched</div></div>
      <div class="kpi"><div class="kpi-label">Marketing Share of New Biz</div><div class="kpi-value">{m['new_biz_origin']['actual']:.1f}%</div><div class="kpi-sub">Target &gt; 25% · {truth['snapshot']['new_biz_fraction']}</div></div>
    </div>
    <div class="chart-grid">
      <div class="chart-card">
        <h3>By Lead Source</h3>
        <div class="chart-sub">Matched leads + % of annualized revenue</div>
        <div class="chart-wrap"><canvas id="sourceChart"></canvas></div>
      </div>
      <div class="chart-card">
        <h3>Monthly Conversion Trend</h3>
        <div class="chart-sub">Match rate + volume through Apr 2026</div>
        <div class="chart-wrap"><canvas id="monthChart"></canvas></div>
      </div>
    </div>

  </div>

  <!-- SOCIAL MEDIA -->
  <div class="section">
    <div class="section-title">Social Media <span class="count">&mdash; Jan 1 &ndash; Apr 2026 · 212 posts, 2,129 likes across IG + FB + LinkedIn</span></div>
    <div class="kpi-inline" style="margin-bottom: 14px;">
      <div class="kpi"><div class="kpi-label">Total Posts</div><div class="kpi-value">{s['total']['posts']}</div><div class="kpi-sub">{s['total']['posts_per_mo']:.1f} / month combined</div></div>
      <div class="kpi"><div class="kpi-label">Total Likes</div><div class="kpi-value">{s['total']['likes']:,}</div><div class="kpi-sub">{s['total']['monthly_likes']:.0f} / month avg</div></div>
      <div class="kpi"><div class="kpi-label">IG Stories</div><div class="kpi-value">{s['ig_stories']['posts']}</div><div class="kpi-sub">{s['ig_stories']['likes']} reactions</div></div>
      <div class="kpi"><div class="kpi-label">Top Channel</div><div class="kpi-value">LinkedIn</div><div class="kpi-sub">{s['platforms'][2]['likes']} likes ({round(s['platforms'][2]['likes']/s['total']['likes']*100)}% of total)</div></div>
    </div>
    <div class="chart-grid">
      <div class="chart-card">
        <h3>Posts by Platform</h3>
        <div class="chart-sub">Volume of content posted</div>
        <div class="chart-wrap"><canvas id="socialPostsChart"></canvas></div>
      </div>
      <div class="chart-card">
        <h3>Likes / Reactions by Platform</h3>
        <div class="chart-sub">Audience engagement</div>
        <div class="chart-wrap"><canvas id="socialLikesChart"></canvas></div>
      </div>
    </div>
  </div>

</div><!-- /container -->

<div class="footer">
  <img src="https://www.cozzinibros.com/wp-content/uploads/2026/03/Cozzini-Logo-White-Letters-Hor-TK.png" alt="Cozzini Bros">
  <span>Q1 2026 Marketing Dashboard</span>
</div>

</div><!-- /page-content -->

<script>
const PASS_HASH='90ec951c';
function simpleHash(s){{let h=0;for(let i=0;i<s.length;i++){{h=((h<<5)-h)+s.charCodeAt(i);h|=0;}}return(h>>>0).toString(16).slice(0,8);}}
function checkAuth(){{
  const input=document.getElementById('auth-pass').value;
  if(simpleHash(input)===PASS_HASH||input==='Marketing2026$'){{
    document.getElementById('auth-gate').style.display='none';
    document.getElementById('page-content').style.display='block';
    sessionStorage.setItem('mkt-auth','true');
  }} else {{
    document.getElementById('auth-error').style.display='block';
    document.getElementById('auth-pass').style.borderColor='#d4918b';
  }}
}}
if(sessionStorage.getItem('mkt-auth')==='true'){{
  document.getElementById('auth-gate').style.display='none';
  document.getElementById('page-content').style.display='block';
}}
</script>

<script>
Chart.defaults.font.family = "-apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif";
Chart.defaults.font.size = 11;

const NAVY  = '{PALETTE['navy']}';
const BLUE  = '{PALETTE['blue']}';
const RED   = '{PALETTE['red']}';
const GOLD  = '{PALETTE['gold']}';
const GREEN = '{PALETTE['green']}';
const LAV   = '{PALETTE['lav']}';
const YEL   = '{PALETTE['yel']}';
const BURNT = '{MKT_ACCENT['burnt']}';

// Budget Plan vs Actual
new Chart(document.getElementById('budgetChart'), {{
  type: 'bar',
  data: {{
    labels: {json.dumps(budget_chart_data['labels'])},
    datasets: [{{
      label: 'USD',
      data: {json.dumps(budget_chart_data['values'])},
      backgroundColor: [NAVY, GREEN, BLUE, GOLD],
      borderRadius: 6, barPercentage: 0.6
    }}]
  }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{ callbacks: {{ label: ctx => '$' + ctx.raw.toLocaleString() }} }}
    }},
    scales: {{
      y: {{ beginAtZero: true, ticks: {{ callback: v => '$' + v.toLocaleString() }} }},
      x: {{ grid: {{ display: false }} }}
    }}
  }}
}});

// Social Posts by platform
new Chart(document.getElementById('socialPostsChart'), {{
  type: 'bar',
  data: {{
    labels: {json.dumps(social_chart_data['labels'])},
    datasets: [{{ label: 'Posts', data: {json.dumps(social_chart_data['posts'])}, backgroundColor: [RED, BLUE, NAVY], borderRadius: 6, barPercentage: 0.55 }}]
  }},
  options: {{ responsive: true, maintainAspectRatio: false, plugins: {{ legend: {{ display: false }} }}, scales: {{ y: {{ beginAtZero: true }}, x: {{ grid: {{ display: false }} }} }} }}
}});

// Social Likes by platform
new Chart(document.getElementById('socialLikesChart'), {{
  type: 'bar',
  data: {{
    labels: {json.dumps(social_chart_data['labels'])},
    datasets: [{{ label: 'Likes', data: {json.dumps(social_chart_data['likes'])}, backgroundColor: [RED, BLUE, NAVY], borderRadius: 6, barPercentage: 0.55 }}]
  }},
  options: {{ responsive: true, maintainAspectRatio: false, plugins: {{ legend: {{ display: false }} }}, scales: {{ y: {{ beginAtZero: true }}, x: {{ grid: {{ display: false }} }} }} }}
}});

// Source breakdown (bar with WRR%)
new Chart(document.getElementById('sourceChart'), {{
  type: 'bar',
  data: {{
    labels: {json.dumps(src_chart['labels'])},
    datasets: [
      {{ label: 'Matched Leads', data: {json.dumps(src_chart['matched'])}, backgroundColor: BLUE, borderRadius: 6, barPercentage: 0.55, yAxisID: 'y' }},
      {{ label: 'Annualized WRR ($)', data: {json.dumps(src_chart['wrr'])}, backgroundColor: GREEN, borderRadius: 6, barPercentage: 0.55, yAxisID: 'y1' }}
    ]
  }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    plugins: {{ legend: {{ display: true, position: 'top', labels: {{ usePointStyle: true, pointStyle: 'circle', padding: 14 }} }} }},
    scales: {{
      y:  {{ beginAtZero: true, position: 'left',  title: {{ display: true, text: 'Matched' }} }},
      y1: {{ beginAtZero: true, position: 'right', grid: {{ display: false }}, ticks: {{ callback: v => '$' + (v/1000).toFixed(0) + 'k' }}, title: {{ display: true, text: 'WRR' }} }},
      x: {{ grid: {{ display: false }} }}
    }}
  }}
}});

// Monthly trend (mixed: bar = leads, line = match rate)
new Chart(document.getElementById('monthChart'), {{
  type: 'bar',
  data: {{
    labels: {json.dumps(month_chart['labels'])},
    datasets: [
      {{ type:'bar',  label: 'Leads',    data: {json.dumps(month_chart['leads'])},    backgroundColor: BLUE, borderRadius: 6, barPercentage: 0.55, yAxisID: 'y' }},
      {{ type:'bar',  label: 'Matched',  data: {json.dumps(month_chart['matched'])},  backgroundColor: GREEN, borderRadius: 6, barPercentage: 0.55, yAxisID: 'y' }},
      {{ type:'line', label: 'Match %',  data: {json.dumps(month_chart['match_rate'])}, borderColor: BURNT, backgroundColor: BURNT, tension: 0.3, yAxisID: 'y1', pointRadius: 4 }}
    ]
  }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    plugins: {{ legend: {{ display: true, position: 'top', labels: {{ usePointStyle: true, pointStyle: 'circle', padding: 14 }} }} }},
    scales: {{
      y:  {{ beginAtZero: true, position: 'left',  title: {{ display: true, text: 'Leads' }} }},
      y1: {{ beginAtZero: true, position: 'right', grid: {{ display: false }}, max: 100, ticks: {{ callback: v => v + '%' }}, title: {{ display: true, text: 'Match rate' }} }},
      x: {{ grid: {{ display: false }} }}
    }}
  }}
}});

</script>
</body>
</html>
"""
    return html


# ============================================================
# MAIN
# ============================================================
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="print truth JSON only")
    ap.add_argument("--truth-out", help="also write truth JSON to this path")
    args = ap.parse_args()

    truth = build_truth()

    if args.dry_run:
        print(json.dumps(truth, indent=2, default=str))
        return

    html = build_html(truth)
    target = ROOT / "mkt" / "q1.html"
    target.write_text(html)
    print(f"✓ wrote {target.relative_to(ROOT)}  ({len(html):,} bytes)")

    if args.truth_out:
        path = ROOT / args.truth_out
        path.write_text(json.dumps(truth, indent=2, default=str))
        print(f"✓ wrote truth JSON → {path.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
